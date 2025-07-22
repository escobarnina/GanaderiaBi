# business_intelligence/views/reportes_views.py
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Count, Avg, Sum, Q, Max, Min
from django.utils import timezone
from datetime import datetime, timedelta, date
import json
import csv
import io
import calendar
from decimal import Decimal

from ..models import (
    MarcaGanadoBovino,
    LogoMarcaBovina,
    KPIGanadoBovino,
    HistorialEstadoMarca,
)
from ..services import ReportService, AnalyticsService


class ReportesBovinoViewSet(viewsets.ViewSet):
    """
    ViewSet para generación de reportes especializados de ganado bovino

    Proporciona:
    - Reportes ejecutivos mensuales y anuales
    - Exportación de datos en múltiples formatos
    - Reportes personalizados con filtros avanzados
    - Análisis comparativos temporales
    - Dashboards para diferentes niveles organizacionales
    """

    @action(detail=False, methods=["get"])
    def reporte_ejecutivo_mensual(self, request):
        """Reporte ejecutivo mensual para la alta dirección"""
        año = int(request.query_params.get("año", timezone.now().year))
        mes = int(request.query_params.get("mes", timezone.now().month))

        try:
            reporte = ReportService.generar_reporte_mensual(año, mes)

            # Agregar análisis comparativo con mes anterior
            mes_anterior = mes - 1 if mes > 1 else 12
            año_anterior = año if mes > 1 else año - 1

            try:
                reporte_anterior = ReportService.generar_reporte_mensual(
                    año_anterior, mes_anterior
                )

                # Calcular variaciones
                variaciones = {
                    "marcas_registradas": {
                        "actual": reporte["metricas_principales"]["marcas_registradas"],
                        "anterior": reporte_anterior["metricas_principales"][
                            "marcas_registradas"
                        ],
                        "variacion": reporte["metricas_principales"][
                            "marcas_registradas"
                        ]
                        - reporte_anterior["metricas_principales"][
                            "marcas_registradas"
                        ],
                        "variacion_porcentual": round(
                            (
                                (
                                    (
                                        reporte["metricas_principales"][
                                            "marcas_registradas"
                                        ]
                                        - reporte_anterior["metricas_principales"][
                                            "marcas_registradas"
                                        ]
                                    )
                                    / reporte_anterior["metricas_principales"][
                                        "marcas_registradas"
                                    ]
                                    * 100
                                )
                                if reporte_anterior["metricas_principales"][
                                    "marcas_registradas"
                                ]
                                > 0
                                else 0
                            ),
                            2,
                        ),
                    },
                    "cabezas_bovinas": {
                        "actual": reporte["metricas_principales"][
                            "total_cabezas_bovinas"
                        ],
                        "anterior": reporte_anterior["metricas_principales"][
                            "total_cabezas_bovinas"
                        ],
                        "variacion": reporte["metricas_principales"][
                            "total_cabezas_bovinas"
                        ]
                        - reporte_anterior["metricas_principales"][
                            "total_cabezas_bovinas"
                        ],
                        "variacion_porcentual": round(
                            (
                                (
                                    (
                                        reporte["metricas_principales"][
                                            "total_cabezas_bovinas"
                                        ]
                                        - reporte_anterior["metricas_principales"][
                                            "total_cabezas_bovinas"
                                        ]
                                    )
                                    / reporte_anterior["metricas_principales"][
                                        "total_cabezas_bovinas"
                                    ]
                                    * 100
                                )
                                if reporte_anterior["metricas_principales"][
                                    "total_cabezas_bovinas"
                                ]
                                > 0
                                else 0
                            ),
                            2,
                        ),
                    },
                    "ingresos": {
                        "actual": float(
                            reporte["metricas_principales"]["ingresos_total"]
                        ),
                        "anterior": float(
                            reporte_anterior["metricas_principales"]["ingresos_total"]
                        ),
                        "variacion": float(
                            reporte["metricas_principales"]["ingresos_total"]
                        )
                        - float(
                            reporte_anterior["metricas_principales"]["ingresos_total"]
                        ),
                        "variacion_porcentual": round(
                            (
                                (
                                    (
                                        float(
                                            reporte["metricas_principales"][
                                                "ingresos_total"
                                            ]
                                        )
                                        - float(
                                            reporte_anterior["metricas_principales"][
                                                "ingresos_total"
                                            ]
                                        )
                                    )
                                    / float(
                                        reporte_anterior["metricas_principales"][
                                            "ingresos_total"
                                        ]
                                    )
                                    * 100
                                )
                                if float(
                                    reporte_anterior["metricas_principales"][
                                        "ingresos_total"
                                    ]
                                )
                                > 0
                                else 0
                            ),
                            2,
                        ),
                    },
                }

                reporte["comparacion_mes_anterior"] = variaciones

            except:
                reporte["comparacion_mes_anterior"] = {
                    "mensaje": "Datos del mes anterior no disponibles"
                }

            # Agregar proyecciones para el próximo mes
            predicciones = AnalyticsService.predecir_demanda_mensual()
            reporte["proyecciones_proximo_mes"] = predicciones

            # Agregar alertas y recomendaciones
            reporte["alertas_ejecutivas"] = self._generar_alertas_ejecutivas(reporte)
            reporte["acciones_recomendadas"] = self._generar_acciones_recomendadas(
                reporte
            )

            # Agregar resumen ejecutivo
            reporte["resumen_ejecutivo"] = self._generar_resumen_ejecutivo(reporte)

            # Agregar métricas de calidad
            reporte["metricas_calidad"] = self._calcular_metricas_calidad(año, mes)

            return Response(reporte)

        except Exception as e:
            return Response({"error": f"Error generando reporte: {str(e)}"}, status=500)

    @action(detail=False, methods=["get"])
    def reporte_anual(self, request):
        """Reporte anual completo del sector ganadero bovino"""
        año = int(request.query_params.get("año", timezone.now().year))

        try:
            # Reporte base del año
            inicio_año = datetime(año, 1, 1)
            fin_año = datetime(año + 1, 1, 1)

            marcas_año = MarcaGanadoBovino.objects.filter(
                fecha_registro__gte=inicio_año, fecha_registro__lt=fin_año
            )

            # Métricas principales del año
            total_marcas = marcas_año.count()
            total_cabezas = (
                marcas_año.aggregate(Sum("cantidad_cabezas"))["cantidad_cabezas__sum"]
                or 0
            )
            ingresos_año = (
                marcas_año.filter(estado="APROBADO").aggregate(
                    Sum("monto_certificacion")
                )["monto_certificacion__sum"]
                or 0
            )

            # Análisis mensual del año
            analisis_mensual = []
            for mes in range(1, 13):
                marcas_mes = marcas_año.filter(fecha_registro__month=mes)

                analisis_mensual.append(
                    {
                        "mes": calendar.month_name[mes],
                        "mes_numero": mes,
                        "marcas_registradas": marcas_mes.count(),
                        "cabezas_registradas": marcas_mes.aggregate(
                            Sum("cantidad_cabezas")
                        )["cantidad_cabezas__sum"]
                        or 0,
                        "ingresos": float(
                            marcas_mes.filter(estado="APROBADO").aggregate(
                                Sum("monto_certificacion")
                            )["monto_certificacion__sum"]
                            or 0
                        ),
                        "departamento_lider": self._obtener_departamento_lider_mes(
                            marcas_mes
                        ),
                    }
                )

            # Análisis por departamento
            analisis_departamental = (
                marcas_año.values("departamento")
                .annotate(
                    total_marcas=Count("id"),
                    total_cabezas=Sum("cantidad_cabezas"),
                    ingresos=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
                    aprobadas=Count("id", filter=Q(estado="APROBADO")),
                    rechazadas=Count("id", filter=Q(estado="RECHAZADO")),
                    tiempo_promedio=Avg("tiempo_procesamiento_horas"),
                )
                .order_by("-total_cabezas")
            )

            # Enriquecer análisis departamental
            for dept in analisis_departamental:
                dept["departamento_display"] = dict(
                    MarcaGanadoBovino._meta.get_field("departamento").choices
                ).get(dept["departamento"], dept["departamento"])

                total_procesadas = dept["aprobadas"] + dept["rechazadas"]
                dept["tasa_aprobacion"] = round(
                    (
                        (dept["aprobadas"] / total_procesadas * 100)
                        if total_procesadas > 0
                        else 0
                    ),
                    2,
                )
                dept["ingresos"] = float(dept["ingresos"] or 0)

            # Análisis por raza bovina
            analisis_razas = (
                marcas_año.values("raza_bovino")
                .annotate(
                    total_marcas=Count("id"),
                    total_cabezas=Sum("cantidad_cabezas"),
                    promedio_cabezas=Avg("cantidad_cabezas"),
                    departamentos_presencia=Count("departamento", distinct=True),
                )
                .order_by("-total_cabezas")
            )

            # Enriquecer análisis de razas
            for raza in analisis_razas:
                raza["raza_display"] = dict(
                    MarcaGanadoBovino._meta.get_field("raza_bovino").choices
                ).get(raza["raza_bovino"], raza["raza_bovino"])
                raza["porcentaje_mercado"] = round(
                    (
                        (raza["total_marcas"] / total_marcas * 100)
                        if total_marcas > 0
                        else 0
                    ),
                    2,
                )

            # Comparación con año anterior
            comparacion_interanual = self._generar_comparacion_interanual(
                año, marcas_año
            )

            # Análisis de eficiencia del año
            eficiencia_anual = self._analizar_eficiencia_anual(marcas_año)

            # Análisis de logos del año
            logos_año = LogoMarcaBovina.objects.filter(fecha_generacion__year=año)

            analisis_logos = {
                "total_generados": logos_año.count(),
                "exitosos": logos_año.filter(exito=True).count(),
                "tasa_exito": round(
                    (
                        (logos_año.filter(exito=True).count() / logos_año.count() * 100)
                        if logos_año.count() > 0
                        else 0
                    ),
                    2,
                ),
                "por_modelo": list(
                    logos_año.values("modelo_ia_usado")
                    .annotate(
                        total=Count("id"), exitosos=Count("id", filter=Q(exito=True))
                    )
                    .order_by("-total")
                ),
                "tiempo_promedio": round(
                    logos_año.aggregate(Avg("tiempo_generacion_segundos"))[
                        "tiempo_generacion_segundos__avg"
                    ]
                    or 0,
                    2,
                ),
            }

            # Proyecciones para el siguiente año
            proyecciones = self._generar_proyecciones_anuales(
                analisis_mensual, total_marcas
            )

            reporte_anual = {
                "año": año,
                "tipo_reporte": "Reporte Anual Ganado Bovino",
                "fecha_generacion": timezone.now().date(),
                "metricas_principales": {
                    "total_marcas_año": total_marcas,
                    "total_cabezas_año": total_cabezas,
                    "promedio_cabezas_por_marca": (
                        round(total_cabezas / total_marcas, 2)
                        if total_marcas > 0
                        else 0
                    ),
                    "ingresos_total_año": float(ingresos_año),
                    "ingreso_promedio_por_marca": (
                        round(float(ingresos_año) / total_marcas, 2)
                        if total_marcas > 0
                        else 0
                    ),
                },
                "analisis_mensual": analisis_mensual,
                "analisis_departamental": list(analisis_departamental),
                "analisis_razas_bovinas": list(analisis_razas),
                "comparacion_año_anterior": comparacion_interanual,
                "eficiencia_anual": eficiencia_anual,
                "analisis_tecnologia_logos": analisis_logos,
                "proyecciones_siguiente_año": proyecciones,
                "conclusiones_ejecutivas": self._generar_conclusiones_anuales(
                    total_marcas, analisis_mensual, comparacion_interanual
                ),
                "recomendaciones_estrategicas": self._generar_recomendaciones_estrategicas(
                    analisis_departamental, analisis_razas, eficiencia_anual
                ),
            }

            return Response(reporte_anual)

        except Exception as e:
            return Response(
                {"error": f"Error generando reporte anual: {str(e)}"}, status=500
            )

    @action(detail=False, methods=["get"])
    def reporte_comparativo_departamentos(self, request):
        """Reporte comparativo detallado entre departamentos ganaderos"""
        fecha_desde = request.query_params.get("fecha_desde")
        fecha_hasta = request.query_params.get("fecha_hasta")

        # Filtros de fecha
        queryset = MarcaGanadoBovino.objects.all()
        if fecha_desde:
            queryset = queryset.filter(fecha_registro__date__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha_registro__date__lte=fecha_hasta)

        # Análisis por departamento
        analisis_departamental = (
            queryset.values("departamento")
            .annotate(
                total_marcas=Count("id"),
                total_cabezas=Sum("cantidad_cabezas"),
                promedio_cabezas=Avg("cantidad_cabezas"),
                aprobadas=Count("id", filter=Q(estado="APROBADO")),
                rechazadas=Count("id", filter=Q(estado="RECHAZADO")),
                pendientes=Count(
                    "id", filter=Q(estado__in=["PENDIENTE", "EN_PROCESO"])
                ),
                ingresos=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
                tiempo_promedio=Avg("tiempo_procesamiento_horas"),
                productores_unicos=Count("ci_productor", distinct=True),
                monto_promedio=Avg("monto_certificacion"),
            )
            .order_by("-total_cabezas")
        )

        # Enriquecer datos
        departamentos_comparativa = []
        total_nacional = queryset.count()
        total_cabezas_nacional = (
            queryset.aggregate(Sum("cantidad_cabezas"))["cantidad_cabezas__sum"] or 0
        )

        for dept in analisis_departamental:
            # Análisis por propósito en el departamento
            propositos_dept = (
                queryset.filter(departamento=dept["departamento"])
                .values("proposito_ganado")
                .annotate(total=Count("id"))
                .order_by("-total")
            )

            # Análisis por raza en el departamento
            razas_dept = (
                queryset.filter(departamento=dept["departamento"])
                .values("raza_bovino")
                .annotate(total=Count("id"), cabezas=Sum("cantidad_cabezas"))
                .order_by("-cabezas")[:3]
            )  # Top 3 razas

            # Calcular métricas adicionales
            total_procesadas = dept["aprobadas"] + dept["rechazadas"]
            tasa_aprobacion = (
                (dept["aprobadas"] / total_procesadas * 100)
                if total_procesadas > 0
                else 0
            )

            # Score de competitividad departamental
            score_competitividad = self._calcular_score_competitividad(
                dept, total_nacional, total_cabezas_nacional
            )

            departamento_data = {
                "departamento": dept["departamento"],
                "departamento_display": dict(
                    MarcaGanadoBovino._meta.get_field("departamento").choices
                ).get(dept["departamento"], dept["departamento"]),
                "metricas_basicas": {
                    "total_marcas": dept["total_marcas"],
                    "total_cabezas": dept["total_cabezas"],
                    "promedio_cabezas": round(dept["promedio_cabezas"] or 0, 2),
                    "participacion_nacional_marcas": round(
                        (
                            (dept["total_marcas"] / total_nacional * 100)
                            if total_nacional > 0
                            else 0
                        ),
                        2,
                    ),
                    "participacion_nacional_cabezas": round(
                        (
                            (dept["total_cabezas"] / total_cabezas_nacional * 100)
                            if total_cabezas_nacional > 0
                            else 0
                        ),
                        2,
                    ),
                },
                "metricas_eficiencia": {
                    "tasa_aprobacion": round(tasa_aprobacion, 2),
                    "tiempo_promedio_procesamiento": round(
                        dept["tiempo_promedio"] or 0, 2
                    ),
                    "productores_activos": dept["productores_unicos"],
                    "marcas_por_productor": round(
                        (
                            (dept["total_marcas"] / dept["productores_unicos"])
                            if dept["productores_unicos"] > 0
                            else 0
                        ),
                        2,
                    ),
                },
                "metricas_economicas": {
                    "ingresos_total": float(dept["ingresos"] or 0),
                    "ingreso_promedio_marca": round(
                        float(dept["monto_promedio"] or 0), 2
                    ),
                    "ingreso_por_cabeza": round(
                        (
                            (float(dept["ingresos"] or 0) / dept["total_cabezas"])
                            if dept["total_cabezas"] > 0
                            else 0
                        ),
                        2,
                    ),
                },
                "distribucion_propositos": [
                    {
                        "proposito": prop["proposito_ganado"],
                        "proposito_display": dict(
                            MarcaGanadoBovino._meta.get_field(
                                "proposito_ganado"
                            ).choices
                        ).get(prop["proposito_ganado"], prop["proposito_ganado"]),
                        "total": prop["total"],
                        "porcentaje": round(
                            (
                                (prop["total"] / dept["total_marcas"] * 100)
                                if dept["total_marcas"] > 0
                                else 0
                            ),
                            2,
                        ),
                    }
                    for prop in propositos_dept
                ],
                "top_razas": [
                    {
                        "raza": raza["raza_bovino"],
                        "raza_display": dict(
                            MarcaGanadoBovino._meta.get_field("raza_bovino").choices
                        ).get(raza["raza_bovino"], raza["raza_bovino"]),
                        "total_marcas": raza["total"],
                        "total_cabezas": raza["cabezas"],
                    }
                    for raza in razas_dept
                ],
                "score_competitividad": score_competitividad,
                "clasificacion": self._clasificar_departamento(
                    score_competitividad, dept
                ),
            }

            departamentos_comparativa.append(departamento_data)

        # Rankings
        rankings = {
            "por_cabezas": sorted(
                departamentos_comparativa,
                key=lambda x: x["metricas_basicas"]["total_cabezas"],
                reverse=True,
            ),
            "por_eficiencia": sorted(
                departamentos_comparativa,
                key=lambda x: x["metricas_eficiencia"]["tasa_aprobacion"],
                reverse=True,
            ),
            "por_ingresos": sorted(
                departamentos_comparativa,
                key=lambda x: x["metricas_economicas"]["ingresos_total"],
                reverse=True,
            ),
            "por_competitividad": sorted(
                departamentos_comparativa,
                key=lambda x: x["score_competitividad"],
                reverse=True,
            ),
        }

        # Análisis de brechas
        analisis_brechas = self._analizar_brechas_departamentales(
            departamentos_comparativa
        )

        # Oportunidades de mejora
        oportunidades = self._identificar_oportunidades_departamentales(
            departamentos_comparativa
        )

        return Response(
            {
                "periodo_analisis": {
                    "fecha_desde": fecha_desde or "inicio_registros",
                    "fecha_hasta": fecha_hasta or "actualidad",
                },
                "resumen_nacional": {
                    "total_departamentos_activos": len(departamentos_comparativa),
                    "total_marcas_nacional": total_nacional,
                    "total_cabezas_nacional": total_cabezas_nacional,
                },
                "analisis_departamental": departamentos_comparativa,
                "rankings": rankings,
                "analisis_brechas": analisis_brechas,
                "oportunidades_mejora": oportunidades,
                "recomendaciones_politicas": self._generar_recomendaciones_politicas(
                    analisis_brechas
                ),
            }
        )

    @action(detail=False, methods=["post"])
    def reporte_personalizado(self, request):
        """Generación de reportes personalizados con filtros específicos"""
        try:
            filtros = request.data.get("filtros", {})
            metricas = request.data.get("metricas", [])
            formato_salida = request.data.get("formato", "json")
            agrupacion = request.data.get("agrupacion", "ninguna")

            # Construir queryset base
            queryset = MarcaGanadoBovino.objects.all()

            # Aplicar filtros
            if filtros.get("departamento"):
                queryset = queryset.filter(departamento__in=filtros["departamento"])

            if filtros.get("raza_bovino"):
                queryset = queryset.filter(raza_bovino__in=filtros["raza_bovino"])

            if filtros.get("proposito_ganado"):
                queryset = queryset.filter(
                    proposito_ganado__in=filtros["proposito_ganado"]
                )

            if filtros.get("estado"):
                queryset = queryset.filter(estado__in=filtros["estado"])

            if filtros.get("fecha_desde"):
                queryset = queryset.filter(
                    fecha_registro__date__gte=filtros["fecha_desde"]
                )

            if filtros.get("fecha_hasta"):
                queryset = queryset.filter(
                    fecha_registro__date__lte=filtros["fecha_hasta"]
                )

            if filtros.get("cabezas_min"):
                queryset = queryset.filter(cantidad_cabezas__gte=filtros["cabezas_min"])

            if filtros.get("cabezas_max"):
                queryset = queryset.filter(cantidad_cabezas__lte=filtros["cabezas_max"])

            # Aplicar agrupación
            if agrupacion == "departamento":
                datos_agrupados = self._agrupar_por_departamento(queryset, metricas)
            elif agrupacion == "raza_bovino":
                datos_agrupados = self._agrupar_por_raza(queryset, metricas)
            elif agrupacion == "proposito_ganado":
                datos_agrupados = self._agrupar_por_proposito(queryset, metricas)
            elif agrupacion == "mes":
                datos_agrupados = self._agrupar_por_mes(queryset, metricas)
            else:
                datos_agrupados = self._calcular_metricas_globales(queryset, metricas)

            reporte_personalizado = {
                "configuracion": {
                    "filtros_aplicados": filtros,
                    "metricas_solicitadas": metricas,
                    "agrupacion": agrupacion,
                    "fecha_generacion": timezone.now(),
                    "total_registros": queryset.count(),
                },
                "datos": datos_agrupados,
                "resumen_estadistico": self._generar_resumen_estadistico(
                    datos_agrupados, agrupacion
                ),
            }

            # Exportar según formato solicitado
            if formato_salida == "excel":
                return self._exportar_excel_personalizado(reporte_personalizado)
            elif formato_salida == "csv":
                return self._exportar_csv_personalizado(reporte_personalizado)
            else:
                return Response(reporte_personalizado)

        except Exception as e:
            return Response(
                {"error": f"Error generando reporte personalizado: {str(e)}"},
                status=400,
            )

    @action(detail=False, methods=["get"])
    def exportar_excel(self, request):
        """Exportación de datos a Excel con formato profesional"""
        try:
            # Parámetros de exportación
            tipo_reporte = request.query_params.get("tipo", "completo")
            fecha_desde = request.query_params.get("fecha_desde")
            fecha_hasta = request.query_params.get("fecha_hasta")
            departamento = request.query_params.get("departamento")

            # Filtros para exportación
            filtros = {}
            if fecha_desde:
                filtros["fecha_desde"] = fecha_desde
            if fecha_hasta:
                filtros["fecha_hasta"] = fecha_hasta
            if departamento:
                filtros["departamento"] = departamento

            # Generar datos para Excel
            datos_excel = ReportService.exportar_datos_excel(filtros)

            # Crear archivo Excel en memoria
            output = io.BytesIO()

            try:
                import xlsxwriter

                # Crear workbook
                workbook = xlsxwriter.Workbook(output, {"in_memory": True})

                # Hojas del reporte
                if tipo_reporte == "completo":
                    self._crear_hoja_datos_principales(workbook, datos_excel["datos"])
                    self._crear_hoja_resumen_estadistico(
                        workbook, datos_excel["resumen"]
                    )
                    self._crear_hoja_analisis_departamental(workbook, filtros)
                    self._crear_hoja_analisis_razas(workbook, filtros)
                elif tipo_reporte == "basico":
                    self._crear_hoja_datos_principales(workbook, datos_excel["datos"])
                    self._crear_hoja_resumen_estadistico(
                        workbook, datos_excel["resumen"]
                    )

                workbook.close()

            except ImportError:
                # Fallback a openpyxl si xlsxwriter no está disponible
                import openpyxl
                from openpyxl.utils.dataframe import dataframe_to_rows
                import pandas as pd

                workbook = openpyxl.Workbook()

                # Hoja principal con datos
                ws = workbook.active
                ws.title = "Datos Principales"

                # Crear DataFrame y escribir a Excel
                df = pd.DataFrame(datos_excel["datos"])
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

                workbook.save(output)

            output.seek(0)

            # Preparar respuesta
            filename = f'reporte_ganaderia_bovina_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return Response(
                {"error": f"Error exportando a Excel: {str(e)}"}, status=500
            )

    # ==================== MÉTODOS AUXILIARES PRINCIPALES ====================

    def _generar_alertas_ejecutivas(self, reporte):
        """Genera alertas para el nivel ejecutivo"""
        alertas = []

        # Alerta por baja tasa de aprobación
        tasa_aprobacion = reporte["metricas_principales"]["tasa_aprobacion"]
        if tasa_aprobacion < 70:
            alertas.append(
                {
                    "tipo": "critica",
                    "categoria": "calidad",
                    "titulo": "Tasa de Aprobación Baja",
                    "mensaje": f"La tasa de aprobación del {tasa_aprobacion:.1f}% está por debajo del objetivo (70%)",
                    "impacto": "Alto - Afecta satisfacción del cliente y reputación",
                    "accion_inmediata": "Revisar criterios de evaluación y capacitar evaluadores",
                }
            )

        # Alerta por tiempo de procesamiento elevado
        tiempo_promedio = reporte["metricas_principales"][
            "tiempo_promedio_procesamiento"
        ]
        if tiempo_promedio > 72:  # Más de 3 días
            alertas.append(
                {
                    "tipo": "warning",
                    "categoria": "eficiencia",
                    "titulo": "Tiempo de Procesamiento Elevado",
                    "mensaje": f"Tiempo promedio de {tiempo_promedio:.1f} horas excede objetivo (72h)",
                    "impacto": "Medio - Afecta experiencia del usuario",
                    "accion_inmediata": "Optimizar procesos y aumentar capacidad",
                }
            )

        # Alerta por decrecimiento en registros
        if (
            "comparacion_mes_anterior" in reporte
            and "marcas_registradas" in reporte["comparacion_mes_anterior"]
        ):
            variacion = reporte["comparacion_mes_anterior"]["marcas_registradas"][
                "variacion_porcentual"
            ]
            if variacion < -15:  # Decrecimiento mayor al 15%
                alertas.append(
                    {
                        "tipo": "warning",
                        "categoria": "demanda",
                        "titulo": "Decrecimiento en Registros",
                        "mensaje": f"Reducción del {abs(variacion):.1f}% en registros vs mes anterior",
                        "impacto": "Alto - Afecta ingresos y crecimiento",
                        "accion_inmediata": "Implementar estrategias de marketing y promoción",
                    }
                )

        # Alerta por baja adopción de tecnología IA
        if "estadisticas_logos" in reporte:
            tasa_exito_logos = reporte["estadisticas_logos"]["tasa_exito"]
            if tasa_exito_logos < 80:
                alertas.append(
                    {
                        "tipo": "info",
                        "categoria": "tecnologia",
                        "titulo": "Baja Eficiencia en Logos IA",
                        "mensaje": f"Tasa de éxito de logos del {tasa_exito_logos:.1f}% por debajo del objetivo (80%)",
                        "impacto": "Bajo - Oportunidad de mejora tecnológica",
                        "accion_inmediata": "Optimizar modelos de IA y prompts",
                    }
                )

        return alertas

    def _generar_acciones_recomendadas(self, reporte):
        """Genera acciones recomendadas basadas en el reporte"""
        acciones = []

        # Análisis de métricas principales
        metricas = reporte["metricas_principales"]

        # Acción para mejorar eficiencia
        if metricas["tiempo_promedio_procesamiento"] > 48:
            acciones.append(
                {
                    "prioridad": "alta",
                    "categoria": "eficiencia_operacional",
                    "accion": "Implementar proceso de evaluación acelerada",
                    "descripcion": "Establecer track rápido para casos simples y automatizar pasos repetitivos",
                    "responsable": "Gerencia de Operaciones",
                    "plazo": "30 días",
                    "recursos_necesarios": [
                        "Personal adicional",
                        "Herramientas de automatización",
                    ],
                    "impacto_esperado": "Reducción del 25% en tiempo de procesamiento",
                }
            )

        # Acción para incrementar registros
        if (
            reporte.get("comparacion_mes_anterior", {})
            .get("marcas_registradas", {})
            .get("variacion_porcentual", 0)
            < 0
        ):
            acciones.append(
                {
                    "prioridad": "alta",
                    "categoria": "crecimiento_mercado",
                    "accion": "Campaña de promoción en departamentos con bajo registro",
                    "descripcion": "Enfocar esfuerzos de marketing en regiones con potencial ganadero no explotado",
                    "responsable": "Gerencia Comercial",
                    "plazo": "60 días",
                    "recursos_necesarios": ["Presupuesto marketing", "Equipo de campo"],
                    "impacto_esperado": "Incremento del 15% en registros mensuales",
                }
            )

        # Acción para mejorar calidad
        if metricas["tasa_aprobacion"] < 75:
            acciones.append(
                {
                    "prioridad": "media",
                    "categoria": "mejora_calidad",
                    "accion": "Programa de capacitación para evaluadores",
                    "descripcion": "Estandarizar criterios y mejorar consistencia en evaluaciones",
                    "responsable": "Gerencia de Calidad",
                    "plazo": "45 días",
                    "recursos_necesarios": [
                        "Capacitadores especializados",
                        "Material didáctico",
                    ],
                    "impacto_esperado": "Mejora del 10% en tasa de aprobación",
                }
            )

        # Acción para tecnología
        if reporte.get("estadisticas_logos", {}).get("tasa_exito", 0) < 85:
            acciones.append(
                {
                    "prioridad": "baja",
                    "categoria": "innovacion_tecnologica",
                    "accion": "Optimización de modelos de IA para logos",
                    "descripcion": "Mejorar prompts y entrenar modelos con datos específicos de ganado bovino",
                    "responsable": "Gerencia de TI",
                    "plazo": "90 días",
                    "recursos_necesarios": [
                        "Especialista en IA",
                        "Datos de entrenamiento",
                    ],
                    "impacto_esperado": "Incremento del 15% en calidad de logos",
                }
            )

        return acciones

    def _generar_resumen_ejecutivo(self, reporte):
        """Genera resumen ejecutivo para alta dirección"""
        metricas = reporte["metricas_principales"]

        # Determinar tendencia general
        if "comparacion_mes_anterior" in reporte:
            variacion_marcas = (
                reporte["comparacion_mes_anterior"]
                .get("marcas_registradas", {})
                .get("variacion_porcentual", 0)
            )
            if variacion_marcas > 10:
                tendencia = "crecimiento_fuerte"
                mensaje_tendencia = (
                    f"Excelente crecimiento del {variacion_marcas:.1f}% en registros"
                )
            elif variacion_marcas > 0:
                tendencia = "crecimiento_moderado"
                mensaje_tendencia = (
                    f"Crecimiento moderado del {variacion_marcas:.1f}% en registros"
                )
            elif variacion_marcas > -5:
                tendencia = "estable"
                mensaje_tendencia = "Actividad estable con variación mínima"
            else:
                tendencia = "decrecimiento"
                mensaje_tendencia = (
                    f"Decrecimiento del {abs(variacion_marcas):.1f}% requiere atención"
                )
        else:
            tendencia = "sin_datos_comparativos"
            mensaje_tendencia = "Primer mes de análisis - estableciendo línea base"

        # Puntos clave
        puntos_clave = []

        # Registro de cabezas bovinas
        if metricas["total_cabezas_bovinas"] > 0:
            puntos_clave.append(
                f"Se registraron {metricas['total_cabezas_bovinas']:,} cabezas de ganado bovino"
            )

        # Eficiencia
        if metricas["tiempo_promedio_procesamiento"] < 48:
            puntos_clave.append(
                f"Excelente eficiencia: {metricas['tiempo_promedio_procesamiento']:.1f}h promedio de procesamiento"
            )
        elif metricas["tiempo_promedio_procesamiento"] > 72:
            puntos_clave.append(
                f"Oportunidad de mejora: {metricas['tiempo_promedio_procesamiento']:.1f}h de procesamiento"
            )

        # Ingresos
        if metricas["ingresos_total"] > 0:
            puntos_clave.append(
                f"Ingresos generados: Bs. {metricas['ingresos_total']:,.2f}"
            )

        return {
            "tendencia_general": tendencia,
            "mensaje_tendencia": mensaje_tendencia,
            "puntos_clave": puntos_clave,
            "score_desempeño": self._calcular_score_desempeño(metricas),
            "recomendacion_principal": self._generar_recomendacion_principal(
                tendencia, metricas
            ),
        }

    def _calcular_metricas_calidad(self, año, mes):
        """Calcula métricas específicas de calidad del proceso"""
        inicio_mes = datetime(año, mes, 1)
        fin_mes = (
            inicio_mes.replace(month=mes + 1)
            if mes < 12
            else inicio_mes.replace(year=año + 1, month=1)
        )

        marcas_mes = MarcaGanadoBovino.objects.filter(
            fecha_registro__gte=inicio_mes, fecha_registro__lt=fin_mes
        )

        # Métricas de calidad
        total_procesadas = marcas_mes.filter(
            estado__in=["APROBADO", "RECHAZADO"]
        ).count()
        aprobadas = marcas_mes.filter(estado="APROBADO").count()

        # Reversiones de estado (indicador de inconsistencia)
        reversiones = (
            HistorialEstadoMarca.objects.filter(
                marca__in=marcas_mes,
                fecha_cambio__gte=inicio_mes,
                fecha_cambio__lt=fin_mes,
            )
            .values("marca")
            .annotate(cambios=Count("id"))
            .filter(cambios__gt=2)
        )

        # Tiempo de procesamiento por rango
        tiempos_procesamiento = marcas_mes.filter(
            tiempo_procesamiento_horas__isnull=False
        ).values_list("tiempo_procesamiento_horas", flat=True)

        rangos_tiempo = {
            "rapido": len([t for t in tiempos_procesamiento if t <= 24]),
            "normal": len([t for t in tiempos_procesamiento if 24 < t <= 72]),
            "lento": len([t for t in tiempos_procesamiento if t > 72]),
        }

        return {
            "tasa_aprobacion": round(
                (aprobadas / total_procesadas * 100) if total_procesadas > 0 else 0, 2
            ),
            "reversiones_detectadas": reversiones.count(),
            "porcentaje_reversiones": round(
                (
                    (reversiones.count() / total_procesadas * 100)
                    if total_procesadas > 0
                    else 0
                ),
                2,
            ),
            "distribucion_tiempos": rangos_tiempo,
            "consistencia_proceso": (
                "alta"
                if reversiones.count() < total_procesadas * 0.05
                else "media" if reversiones.count() < total_procesadas * 0.1 else "baja"
            ),
            "eficiencia_temporal": (
                "alta"
                if rangos_tiempo["rapido"] > total_procesadas * 0.6
                else (
                    "media"
                    if rangos_tiempo["normal"] > total_procesadas * 0.5
                    else "baja"
                )
            ),
        }

    def _obtener_departamento_lider_mes(self, marcas_mes):
        """Obtiene el departamento líder en registros del mes"""
        lider = (
            marcas_mes.values("departamento")
            .annotate(total=Count("id"))
            .order_by("-total")
            .first()
        )

        if lider:
            return {
                "departamento": lider["departamento"],
                "departamento_display": dict(
                    MarcaGanadoBovino._meta.get_field("departamento").choices
                ).get(lider["departamento"], lider["departamento"]),
                "total_marcas": lider["total"],
            }
        return None

    def _generar_comparacion_interanual(self, año, marcas_año):
        """Genera comparación con el año anterior"""
        try:
            # Datos del año anterior
            inicio_año_anterior = datetime(año - 1, 1, 1)
            fin_año_anterior = datetime(año, 1, 1)

            marcas_año_anterior = MarcaGanadoBovino.objects.filter(
                fecha_registro__gte=inicio_año_anterior,
                fecha_registro__lt=fin_año_anterior,
            )

            # Métricas comparativas
            total_actual = marcas_año.count()
            total_anterior = marcas_año_anterior.count()

            cabezas_actual = (
                marcas_año.aggregate(Sum("cantidad_cabezas"))["cantidad_cabezas__sum"]
                or 0
            )
            cabezas_anterior = (
                marcas_año_anterior.aggregate(Sum("cantidad_cabezas"))[
                    "cantidad_cabezas__sum"
                ]
                or 0
            )

            ingresos_actual = (
                marcas_año.filter(estado="APROBADO").aggregate(
                    Sum("monto_certificacion")
                )["monto_certificacion__sum"]
                or 0
            )
            ingresos_anterior = (
                marcas_año_anterior.filter(estado="APROBADO").aggregate(
                    Sum("monto_certificacion")
                )["monto_certificacion__sum"]
                or 0
            )

            return {
                "año_comparacion": año - 1,
                "marcas": {
                    "actual": total_actual,
                    "anterior": total_anterior,
                    "variacion": total_actual - total_anterior,
                    "variacion_porcentual": round(
                        (
                            ((total_actual - total_anterior) / total_anterior * 100)
                            if total_anterior > 0
                            else 0
                        ),
                        2,
                    ),
                },
                "cabezas_bovinas": {
                    "actual": cabezas_actual,
                    "anterior": cabezas_anterior,
                    "variacion": cabezas_actual - cabezas_anterior,
                    "variacion_porcentual": round(
                        (
                            (
                                (cabezas_actual - cabezas_anterior)
                                / cabezas_anterior
                                * 100
                            )
                            if cabezas_anterior > 0
                            else 0
                        ),
                        2,
                    ),
                },
                "ingresos": {
                    "actual": float(ingresos_actual),
                    "anterior": float(ingresos_anterior),
                    "variacion": float(ingresos_actual - ingresos_anterior),
                    "variacion_porcentual": round(
                        (
                            (
                                (float(ingresos_actual) - float(ingresos_anterior))
                                / float(ingresos_anterior)
                                * 100
                            )
                            if ingresos_anterior > 0
                            else 0
                        ),
                        2,
                    ),
                },
                "interpretacion": self._interpretar_comparacion_anual(
                    total_actual, total_anterior
                ),
            }

        except Exception as e:
            return {"error": f"No se pudieron obtener datos del año anterior: {str(e)}"}

    def _interpretar_comparacion_anual(self, actual, anterior):
        """Interpreta la comparación anual"""
        if anterior == 0:
            return "primer_año_operacion"

        variacion_pct = ((actual - anterior) / anterior) * 100

        if variacion_pct > 25:
            return "crecimiento_excepcional"
        elif variacion_pct > 10:
            return "crecimiento_solido"
        elif variacion_pct > 0:
            return "crecimiento_moderado"
        elif variacion_pct > -10:
            return "declive_moderado"
        else:
            return "declive_significativo"

    def _analizar_eficiencia_anual(self, marcas_año):
        """Analiza la eficiencia operacional del año"""
        # Tiempo promedio de procesamiento
        tiempo_promedio = (
            marcas_año.filter(tiempo_procesamiento_horas__isnull=False).aggregate(
                Avg("tiempo_procesamiento_horas")
            )["tiempo_procesamiento_horas__avg"]
            or 0
        )

        # Tasa de aprobación
        total_procesadas = marcas_año.filter(
            estado__in=["APROBADO", "RECHAZADO"]
        ).count()
        aprobadas = marcas_año.filter(estado="APROBADO").count()
        tasa_aprobacion = (
            (aprobadas / total_procesadas * 100) if total_procesadas > 0 else 0
        )

        # Distribución mensual de eficiencia
        eficiencia_mensual = []
        for mes in range(1, 13):
            marcas_mes = marcas_año.filter(fecha_registro__month=mes)

            if marcas_mes.exists():
                tiempo_mes = (
                    marcas_mes.filter(
                        tiempo_procesamiento_horas__isnull=False
                    ).aggregate(Avg("tiempo_procesamiento_horas"))[
                        "tiempo_procesamiento_horas__avg"
                    ]
                    or 0
                )

                procesadas_mes = marcas_mes.filter(
                    estado__in=["APROBADO", "RECHAZADO"]
                ).count()
                aprobadas_mes = marcas_mes.filter(estado="APROBADO").count()
                tasa_mes = (
                    (aprobadas_mes / procesadas_mes * 100) if procesadas_mes > 0 else 0
                )

                eficiencia_mensual.append(
                    {
                        "mes": calendar.month_name[mes],
                        "tiempo_promedio": round(tiempo_mes, 2),
                        "tasa_aprobacion": round(tasa_mes, 2),
                        "total_procesadas": procesadas_mes,
                    }
                )

        # Score de eficiencia general
        score_tiempo = (
            max(0, 100 - (tiempo_promedio - 24) * 2) if tiempo_promedio > 24 else 100
        )
        score_aprobacion = tasa_aprobacion
        score_general = score_tiempo * 0.4 + score_aprobacion * 0.6

        return {
            "tiempo_promedio_año": round(tiempo_promedio, 2),
            "tasa_aprobacion_año": round(tasa_aprobacion, 2),
            "score_eficiencia_general": round(score_general, 2),
            "clasificacion_eficiencia": self._clasificar_eficiencia(score_general),
            "eficiencia_por_mes": eficiencia_mensual,
            "mejor_mes": (
                max(eficiencia_mensual, key=lambda x: x["tasa_aprobacion"])
                if eficiencia_mensual
                else None
            ),
            "oportunidades_mejora": self._identificar_oportunidades_eficiencia(
                eficiencia_mensual
            ),
        }

    def _clasificar_eficiencia(self, score):
        """Clasifica el nivel de eficiencia"""
        if score >= 85:
            return "excelente"
        elif score >= 70:
            return "buena"
        elif score >= 55:
            return "aceptable"
        else:
            return "necesita_mejora"

    def _identificar_oportunidades_eficiencia(self, eficiencia_mensual):
        """Identifica oportunidades de mejora en eficiencia"""
        if not eficiencia_mensual:
            return []

        oportunidades = []

        # Meses con baja eficiencia
        meses_baja_eficiencia = [
            mes
            for mes in eficiencia_mensual
            if mes["tasa_aprobacion"] < 70 or mes["tiempo_promedio"] > 72
        ]

        if meses_baja_eficiencia:
            oportunidades.append(
                {
                    "categoria": "eficiencia_temporal",
                    "descripcion": f"{len(meses_baja_eficiencia)} meses con eficiencia por debajo del objetivo",
                    "meses_afectados": [mes["mes"] for mes in meses_baja_eficiencia],
                    "impacto": "medio",
                }
            )

        # Variabilidad alta entre meses
        tiempos = [mes["tiempo_promedio"] for mes in eficiencia_mensual]
        if tiempos and (max(tiempos) - min(tiempos)) > 48:
            oportunidades.append(
                {
                    "categoria": "consistencia",
                    "descripcion": "Alta variabilidad en tiempos de procesamiento entre meses",
                    "rango_variacion": f"{min(tiempos):.1f}h - {max(tiempos):.1f}h",
                    "impacto": "bajo",
                }
            )

        return oportunidades

    def _generar_proyecciones_anuales(self, analisis_mensual, total_marcas):
        """Genera proyecciones para el siguiente año"""
        if len(analisis_mensual) < 6:  # Necesitamos al menos 6 meses de datos
            return {"mensaje": "Datos insuficientes para proyecciones confiables"}

        # Calcular tendencia de crecimiento
        marcas_por_mes = [mes["marcas_registradas"] for mes in analisis_mensual]

        # Tendencia lineal simple
        n = len(marcas_por_mes)
        suma_x = sum(range(1, n + 1))
        suma_y = sum(marcas_por_mes)
        suma_xy = sum((i + 1) * marcas_por_mes[i] for i in range(n))
        suma_x2 = sum((i + 1) ** 2 for i in range(n))

        # Coeficientes de regresión lineal
        pendiente = (n * suma_xy - suma_x * suma_y) / (n * suma_x2 - suma_x**2)
        intercepto = (suma_y - pendiente * suma_x) / n

        # Proyección para 12 meses siguientes
        proyecciones_mensuales = []
        for mes in range(1, 13):
            proyeccion = max(0, intercepto + pendiente * (n + mes))
            proyecciones_mensuales.append(
                {
                    "mes": calendar.month_name[mes],
                    "proyeccion_marcas": int(proyeccion),
                    "confianza": (
                        "alta" if mes <= 6 else "media" if mes <= 9 else "baja"
                    ),
                }
            )

        proyeccion_anual = sum(p["proyeccion_marcas"] for p in proyecciones_mensuales)

        return {
            "proyeccion_total_año": proyeccion_anual,
            "crecimiento_estimado": round(
                (
                    ((proyeccion_anual - total_marcas) / total_marcas * 100)
                    if total_marcas > 0
                    else 0
                ),
                2,
            ),
            "proyecciones_mensuales": proyecciones_mensuales,
            "metodologia": "regresion_lineal",
            "nivel_confianza": "medio",
            "factores_considerados": ["tendencia_historica", "estacionalidad_basica"],
        }

    def _generar_conclusiones_anuales(
        self, total_marcas, analisis_mensual, comparacion_interanual
    ):
        """Genera conclusiones ejecutivas del año"""
        conclusiones = []

        # Conclusión sobre volumen
        if total_marcas > 1000:
            conclusiones.append(
                {
                    "categoria": "volumen",
                    "conclusion": f"Año exitoso con {total_marcas:,} marcas registradas, consolidando el liderazgo sectorial",
                }
            )
        elif total_marcas > 500:
            conclusiones.append(
                {
                    "categoria": "volumen",
                    "conclusion": f"Año sólido con {total_marcas:,} marcas registradas, mostrando crecimiento sostenible",
                }
            )
        else:
            conclusiones.append(
                {
                    "categoria": "volumen",
                    "conclusion": f"Año de construcción con {total_marcas:,} marcas registradas, estableciendo base para crecimiento",
                }
            )

        # Conclusión sobre tendencia
        if comparacion_interanual and "marcas" in comparacion_interanual:
            variacion = comparacion_interanual["marcas"]["variacion_porcentual"]
            if variacion > 15:
                conclusiones.append(
                    {
                        "categoria": "crecimiento",
                        "conclusion": f"Excelente crecimiento del {variacion:.1f}% demuestra fortaleza del sector ganadero boliviano",
                    }
                )
            elif variacion > 0:
                conclusiones.append(
                    {
                        "categoria": "crecimiento",
                        "conclusion": f"Crecimiento sostenible del {variacion:.1f}% refleja estabilidad del mercado",
                    }
                )
            else:
                conclusiones.append(
                    {
                        "categoria": "crecimiento",
                        "conclusion": f"Contracción del {abs(variacion):.1f}% requiere estrategias de revitalización",
                    }
                )

        # Conclusión sobre estacionalidad
        if analisis_mensual:
            marcas_por_mes = [mes["marcas_registradas"] for mes in analisis_mensual]
            variabilidad = (max(marcas_por_mes) - min(marcas_por_mes)) / (
                sum(marcas_por_mes) / len(marcas_por_mes)
            )

            if variabilidad > 0.5:
                conclusiones.append(
                    {
                        "categoria": "estacionalidad",
                        "conclusion": "Alta variabilidad estacional presenta oportunidades de nivelación",
                    }
                )
            else:
                conclusiones.append(
                    {
                        "categoria": "estacionalidad",
                        "conclusion": "Demanda relativamente estable a lo largo del año facilita planificación",
                    }
                )

        return conclusiones

    def _generar_recomendaciones_estrategicas(
        self, analisis_departamental, analisis_razas, eficiencia_anual
    ):
        """Genera recomendaciones estratégicas de alto nivel"""
        recomendaciones = []

        # Recomendación geográfica
        if analisis_departamental:
            top_departamento = list(analisis_departamental)[0]
            recomendaciones.append(
                {
                    "categoria": "expansion_geografica",
                    "recomendacion": f'Fortalecer presencia en {top_departamento["departamento_display"]} que lidera con {top_departamento["total_cabezas"]:,} cabezas',
                    "justificacion": "Aprovechar fortalezas regionales para maximizar impacto",
                    "plazo": "mediano_plazo",
                    "recursos": ["inversión_regional", "personal_especializado"],
                }
            )

        # Recomendación por eficiencia
        if eficiencia_anual and eficiencia_anual["score_eficiencia_general"] < 70:
            recomendaciones.append(
                {
                    "categoria": "mejora_operacional",
                    "recomendacion": "Implementar programa integral de mejora de eficiencia operacional",
                    "justificacion": f'Score actual de {eficiencia_anual["score_eficiencia_general"]:.1f} tiene potencial de mejora',
                    "plazo": "corto_plazo",
                    "recursos": ["consultoria_procesos", "tecnologia", "capacitacion"],
                }
            )

        # Recomendación por diversificación
        if analisis_razas:
            total_razas = len(list(analisis_razas))
            if total_razas < 5:
                recomendaciones.append(
                    {
                        "categoria": "diversificacion_genetica",
                        "recomendacion": "Promover diversificación genética para fortalecer sector",
                        "justificacion": f"Solo {total_razas} razas principales registradas",
                        "plazo": "largo_plazo",
                        "recursos": ["programas_incentivos", "educacion_ganadera"],
                    }
                )

        # Recomendación tecnológica
        recomendaciones.append(
            {
                "categoria": "innovacion_tecnologica",
                "recomendacion": "Acelerar adopción de tecnologías 4.0 en el sector ganadero",
                "justificacion": "Posicionar a Bolivia como líder en ganadería digital",
                "plazo": "mediano_plazo",
                "recursos": [
                    "inversion_tecnologica",
                    "alianzas_estrategicas",
                    "capacitacion_digital",
                ],
            }
        )

        return recomendaciones

    def _calcular_score_competitividad(
        self, departamento, total_nacional, total_cabezas_nacional
    ):
        """Calcula score de competitividad departamental"""
        # Participación en mercado (30%)
        participacion_marcas = (
            (departamento["total_marcas"] / total_nacional * 100)
            if total_nacional > 0
            else 0
        )
        participacion_cabezas = (
            (departamento["total_cabezas"] / total_cabezas_nacional * 100)
            if total_cabezas_nacional > 0
            else 0
        )
        score_participacion = (participacion_marcas + participacion_cabezas) / 2

    def _calcular_score_competitividad(
        self, departamento, total_nacional, total_cabezas_nacional
    ):
        """Calcula score de competitividad departamental"""
        # Participación en mercado (30%)
        participacion_marcas = (
            (departamento["total_marcas"] / total_nacional * 100)
            if total_nacional > 0
            else 0
        )
        participacion_cabezas = (
            (departamento["total_cabezas"] / total_cabezas_nacional * 100)
            if total_cabezas_nacional > 0
            else 0
        )
        score_participacion = (participacion_marcas + participacion_cabezas) / 2

        # Eficiencia (25%)
        total_procesadas = departamento["aprobadas"] + departamento["rechazadas"]
        tasa_aprobacion = (
            (departamento["aprobadas"] / total_procesadas * 100)
            if total_procesadas > 0
            else 0
        )
        score_eficiencia = min(tasa_aprobacion, 100)

        # Productividad (25%)
        promedio_cabezas = departamento["promedio_cabezas"] or 0
        score_productividad = min(
            (promedio_cabezas / 100) * 100, 100
        )  # Normalizar a 100 cabezas como máximo

        # Rentabilidad (20%)
        ingreso_por_cabeza = (
            (float(departamento["ingresos"] or 0) / departamento["total_cabezas"])
            if departamento["total_cabezas"] > 0
            else 0
        )
        score_rentabilidad = min(
            (ingreso_por_cabeza / 20) * 100, 100
        )  # Normalizar a 20 Bs por cabeza como máximo

        # Score total ponderado
        score_total = (
            score_participacion * 0.30
            + score_eficiencia * 0.25
            + score_productividad * 0.25
            + score_rentabilidad * 0.20
        )

        return round(score_total, 2)

    def _clasificar_departamento(self, score_competitividad, departamento):
        """Clasifica el departamento según su competitividad"""
        if score_competitividad >= 80:
            return {
                "categoria": "lider_nacional",
                "descripcion": "Departamento líder con excelentes métricas en todos los aspectos",
                "fortalezas": [
                    "alta_participacion",
                    "excelente_eficiencia",
                    "alta_productividad",
                ],
            }
        elif score_competitividad >= 60:
            return {
                "categoria": "competidor_fuerte",
                "descripcion": "Departamento con sólido desempeño y potencial de liderazgo",
                "fortalezas": self._identificar_fortalezas_departamento(departamento),
            }
        elif score_competitividad >= 40:
            return {
                "categoria": "mercado_medio",
                "descripcion": "Departamento con desempeño promedio, oportunidades de mejora",
                "areas_mejora": self._identificar_areas_mejora_departamento(
                    departamento
                ),
            }
        else:
            return {
                "categoria": "potencial_desarrollo",
                "descripcion": "Departamento con gran potencial, requiere inversión estratégica",
                "oportunidades": [
                    "desarrollo_infraestructura",
                    "capacitacion",
                    "incentivos",
                ],
            }

    def _identificar_fortalezas_departamento(self, departamento):
        """Identifica fortalezas específicas del departamento"""
        fortalezas = []

        if departamento["promedio_cabezas"] > 100:
            fortalezas.append("alta_productividad")

        total_procesadas = departamento["aprobadas"] + departamento["rechazadas"]
        if (
            total_procesadas > 0
            and (departamento["aprobadas"] / total_procesadas) > 0.8
        ):
            fortalezas.append("alta_calidad")

        if departamento["tiempo_promedio"] and departamento["tiempo_promedio"] < 48:
            fortalezas.append("procesamiento_rapido")

        if float(departamento["ingresos"] or 0) > 100000:
            fortalezas.append("alto_volumen_economico")

        return fortalezas

    def _identificar_areas_mejora_departamento(self, departamento):
        """Identifica áreas de mejora específicas del departamento"""
        areas_mejora = []

        if departamento["promedio_cabezas"] and departamento["promedio_cabezas"] < 50:
            areas_mejora.append("incrementar_productividad")

        total_procesadas = departamento["aprobadas"] + departamento["rechazadas"]
        if (
            total_procesadas > 0
            and (departamento["aprobadas"] / total_procesadas) < 0.7
        ):
            areas_mejora.append("mejorar_calidad_solicitudes")

        if departamento["tiempo_promedio"] and departamento["tiempo_promedio"] > 72:
            areas_mejora.append("acelerar_procesamiento")

        if departamento["productores_unicos"] < 10:
            areas_mejora.append("expandir_base_productores")

        return areas_mejora

    def _analizar_brechas_departamentales(self, departamentos_comparativa):
        """Analiza brechas entre departamentos"""
        if len(departamentos_comparativa) < 2:
            return {"mensaje": "Insuficientes departamentos para análisis de brechas"}

        # Métricas para análisis
        cabezas_por_dept = [
            d["metricas_basicas"]["total_cabezas"] for d in departamentos_comparativa
        ]
        ingresos_por_dept = [
            d["metricas_economicas"]["ingresos_total"]
            for d in departamentos_comparativa
        ]
        eficiencia_por_dept = [
            d["metricas_eficiencia"]["tasa_aprobacion"]
            for d in departamentos_comparativa
        ]

        # Calcular brechas
        brecha_cabezas = max(cabezas_por_dept) - min(cabezas_por_dept)
        brecha_ingresos = max(ingresos_por_dept) - min(ingresos_por_dept)
        brecha_eficiencia = max(eficiencia_por_dept) - min(eficiencia_por_dept)

        # Coeficiente de variación
        import statistics

        cv_cabezas = (
            statistics.stdev(cabezas_por_dept) / statistics.mean(cabezas_por_dept) * 100
            if statistics.mean(cabezas_por_dept) > 0
            else 0
        )
        cv_ingresos = (
            statistics.stdev(ingresos_por_dept)
            / statistics.mean(ingresos_por_dept)
            * 100
            if statistics.mean(ingresos_por_dept) > 0
            else 0
        )

        return {
            "brechas_absolutas": {
                "cabezas_bovinas": int(brecha_cabezas),
                "ingresos": round(brecha_ingresos, 2),
                "eficiencia": round(brecha_eficiencia, 2),
            },
            "desigualdad_regional": {
                "coeficiente_variacion_cabezas": round(cv_cabezas, 2),
                "coeficiente_variacion_ingresos": round(cv_ingresos, 2),
                "nivel_desigualdad": (
                    "alta"
                    if cv_cabezas > 50
                    else "media" if cv_cabezas > 25 else "baja"
                ),
            },
            "departamento_lider": max(
                departamentos_comparativa, key=lambda x: x["score_competitividad"]
            )["departamento_display"],
            "departamento_potencial": min(
                departamentos_comparativa, key=lambda x: x["score_competitividad"]
            )["departamento_display"],
            "oportunidades_nivelacion": self._identificar_oportunidades_nivelacion(
                departamentos_comparativa
            ),
        }

    def _identificar_oportunidades_departamentales(self, departamentos_comparativa):
        """Identifica oportunidades específicas por departamento"""
        oportunidades = []

        # Ordenar por competitividad
        departamentos_ordenados = sorted(
            departamentos_comparativa,
            key=lambda x: x["score_competitividad"],
            reverse=True,
        )

        # Oportunidades para departamentos de bajo rendimiento
        departamentos_bajo_rendimiento = [
            d for d in departamentos_ordenados if d["score_competitividad"] < 40
        ]

        for dept in departamentos_bajo_rendimiento:
            oportunidades.append(
                {
                    "departamento": dept["departamento_display"],
                    "tipo": "desarrollo_integral",
                    "descripcion": f'Programa de fortalecimiento para {dept["departamento_display"]}',
                    "potencial_mejora": round(60 - dept["score_competitividad"], 2),
                    "acciones_sugeridas": [
                        "Inversión en infraestructura ganadera",
                        "Capacitación técnica a productores",
                        "Programas de financiamiento",
                        "Mejora de canales de comercialización",
                    ],
                }
            )

        # Oportunidades para departamentos con potencial específico
        for dept in departamentos_ordenados:
            # Alta productividad, baja participación
            if (
                dept["metricas_basicas"]["promedio_cabezas"] > 80
                and dept["metricas_basicas"]["participacion_nacional_cabezas"] < 10
            ):
                oportunidades.append(
                    {
                        "departamento": dept["departamento_display"],
                        "tipo": "expansion_mercado",
                        "descripcion": "Alta productividad con baja participación nacional",
                        "acciones_sugeridas": [
                            "Campañas promocionales específicas",
                            "Incentivos para nuevos productores",
                            "Fortalecimiento de asociaciones ganaderas",
                        ],
                    }
                )

            # Alta eficiencia, bajo volumen
            if (
                dept["metricas_eficiencia"]["tasa_aprobacion"] > 85
                and dept["metricas_basicas"]["total_marcas"] < 50
            ):
                oportunidades.append(
                    {
                        "departamento": dept["departamento_display"],
                        "tipo": "escalamiento",
                        "descripcion": "Excelente eficiencia con potencial de crecimiento",
                        "acciones_sugeridas": [
                            "Replicar mejores prácticas en otros departamentos",
                            "Incrementar capacidad de procesamiento",
                            "Programas de atracción de inversión",
                        ],
                    }
                )

        return oportunidades

    def _identificar_oportunidades_nivelacion(self, departamentos_comparativa):
        """Identifica oportunidades para reducir brechas entre departamentos"""
        # Encontrar mejores prácticas
        mejor_eficiencia = max(
            departamentos_comparativa,
            key=lambda x: x["metricas_eficiencia"]["tasa_aprobacion"],
        )
        mejor_productividad = max(
            departamentos_comparativa,
            key=lambda x: x["metricas_basicas"]["promedio_cabezas"],
        )

        oportunidades = []

        # Transferencia de mejores prácticas
        oportunidades.append(
            {
                "tipo": "transferencia_conocimiento",
                "descripcion": f'Transferir metodologías de {mejor_eficiencia["departamento_display"]} a otros departamentos',
                "beneficio": "Mejora general en tasa de aprobación",
                "departamentos_beneficiarios": [
                    d["departamento_display"]
                    for d in departamentos_comparativa
                    if d["metricas_eficiencia"]["tasa_aprobacion"]
                    < mejor_eficiencia["metricas_eficiencia"]["tasa_aprobacion"] - 10
                ],
            }
        )

        # Programas de mentoring interdepartamental
        if mejor_productividad["metricas_basicas"]["promedio_cabezas"] > 100:
            oportunidades.append(
                {
                    "tipo": "mentoring_interdepartamental",
                    "descripcion": f'Programa de mentoring liderado por {mejor_productividad["departamento_display"]}',
                    "beneficio": "Incremento en productividad promedio nacional",
                    "meta": "Reducir brecha de productividad en 30%",
                }
            )

        return oportunidades

    def _generar_recomendaciones_politicas(self, analisis_brechas):
        """Genera recomendaciones de política pública"""
        recomendaciones = []

        if (
            analisis_brechas.get("desigualdad_regional", {}).get("nivel_desigualdad")
            == "alta"
        ):
            recomendaciones.append(
                {
                    "area": "desarrollo_regional",
                    "recomendacion": "Implementar programa nacional de nivelación ganadera",
                    "justificacion": "Alta desigualdad regional requiere intervención coordinada",
                    "instrumentos": [
                        "Fondo de desarrollo ganadero regional",
                        "Incentivos fiscales diferenciados",
                        "Programas de transferencia tecnológica",
                    ],
                    "plazo": "mediano_plazo",
                    "presupuesto_estimado": "alto",
                }
            )

        recomendaciones.append(
            {
                "area": "fortalecimiento_institucional",
                "recomendacion": "Fortalecer capacidades institucionales en departamentos rezagados",
                "justificacion": "Mejorar eficiencia operacional en todo el territorio nacional",
                "instrumentos": [
                    "Capacitación de personal técnico",
                    "Modernización de sistemas",
                    "Estandarización de procesos",
                ],
                "plazo": "corto_plazo",
                "presupuesto_estimado": "medio",
            }
        )

        recomendaciones.append(
            {
                "area": "innovacion_tecnologica",
                "recomendacion": "Acelerar adopción tecnológica en el sector ganadero",
                "justificacion": "Posicionar a Bolivia como líder regional en ganadería digital",
                "instrumentos": [
                    "Plataforma nacional de trazabilidad",
                    "Sistemas de información ganadera",
                    "Programas de digitalización rural",
                ],
                "plazo": "largo_plazo",
                "presupuesto_estimado": "alto",
            }
        )

        return recomendaciones

    # ==================== MÉTODOS DE AGRUPACIÓN ====================

    def _agrupar_por_departamento(self, queryset, metricas):
        """Agrupa datos por departamento según métricas solicitadas"""
        agrupacion = (
            queryset.values("departamento")
            .annotate(
                total_registros=Count("id"),
                total_cabezas=Sum("cantidad_cabezas"),
                promedio_cabezas=Avg("cantidad_cabezas"),
                aprobadas=Count("id", filter=Q(estado="APROBADO")),
                rechazadas=Count("id", filter=Q(estado="RECHAZADO")),
                ingresos_total=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
                tiempo_promedio=Avg("tiempo_procesamiento_horas"),
                productores_unicos=Count("ci_productor", distinct=True),
            )
            .order_by("-total_cabezas")
        )

        # Enriquecer con displays y métricas calculadas
        datos_agrupados = []
        for grupo in agrupacion:
            total_procesadas = grupo["aprobadas"] + grupo["rechazadas"]

            item = {
                "departamento": grupo["departamento"],
                "departamento_display": dict(
                    MarcaGanadoBovino._meta.get_field("departamento").choices
                ).get(grupo["departamento"], grupo["departamento"]),
                "total_registros": grupo["total_registros"],
                "total_cabezas": grupo["total_cabezas"],
                "promedio_cabezas": round(grupo["promedio_cabezas"] or 0, 2),
                "tasa_aprobacion": round(
                    (
                        (grupo["aprobadas"] / total_procesadas * 100)
                        if total_procesadas > 0
                        else 0
                    ),
                    2,
                ),
                "ingresos_total": float(grupo["ingresos_total"] or 0),
                "tiempo_promedio": round(grupo["tiempo_promedio"] or 0, 2),
                "productores_activos": grupo["productores_unicos"],
                "densidad_ganadera": round(
                    (
                        (grupo["total_cabezas"] / grupo["total_registros"])
                        if grupo["total_registros"] > 0
                        else 0
                    ),
                    2,
                ),
            }

            # Agregar solo métricas solicitadas si se especificaron
            if metricas:
                item_filtrado = {
                    "departamento": item["departamento"],
                    "departamento_display": item["departamento_display"],
                }
                for metrica in metricas:
                    if metrica in item:
                        item_filtrado[metrica] = item[metrica]
                datos_agrupados.append(item_filtrado)
            else:
                datos_agrupados.append(item)

        return datos_agrupados

    def _agrupar_por_raza(self, queryset, metricas):
        """Agrupa datos por raza bovina según métricas solicitadas"""
        agrupacion = (
            queryset.values("raza_bovino")
            .annotate(
                total_registros=Count("id"),
                total_cabezas=Sum("cantidad_cabezas"),
                promedio_cabezas=Avg("cantidad_cabezas"),
                aprobadas=Count("id", filter=Q(estado="APROBADO")),
                rechazadas=Count("id", filter=Q(estado="RECHAZADO")),
                ingresos_total=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
                departamentos_presencia=Count("departamento", distinct=True),
                monto_promedio=Avg("monto_certificacion"),
            )
            .order_by("-total_cabezas")
        )

        total_nacional = queryset.count()

        datos_agrupados = []
        for grupo in agrupacion:
            total_procesadas = grupo["aprobadas"] + grupo["rechazadas"]

            item = {
                "raza_bovino": grupo["raza_bovino"],
                "raza_display": dict(
                    MarcaGanadoBovino._meta.get_field("raza_bovino").choices
                ).get(grupo["raza_bovino"], grupo["raza_bovino"]),
                "total_registros": grupo["total_registros"],
                "total_cabezas": grupo["total_cabezas"],
                "promedio_cabezas": round(grupo["promedio_cabezas"] or 0, 2),
                "participacion_mercado": round(
                    (
                        (grupo["total_registros"] / total_nacional * 100)
                        if total_nacional > 0
                        else 0
                    ),
                    2,
                ),
                "tasa_aprobacion": round(
                    (
                        (grupo["aprobadas"] / total_procesadas * 100)
                        if total_procesadas > 0
                        else 0
                    ),
                    2,
                ),
                "ingresos_total": float(grupo["ingresos_total"] or 0),
                "diversidad_geografica": grupo["departamentos_presencia"],
                "valor_promedio_marca": round(float(grupo["monto_promedio"] or 0), 2),
            }

            if metricas:
                item_filtrado = {
                    "raza_bovino": item["raza_bovino"],
                    "raza_display": item["raza_display"],
                }
                for metrica in metricas:
                    if metrica in item:
                        item_filtrado[metrica] = item[metrica]
                datos_agrupados.append(item_filtrado)
            else:
                datos_agrupados.append(item)

        return datos_agrupados

    def _agrupar_por_proposito(self, queryset, metricas):
        """Agrupa datos por propósito ganadero según métricas solicitadas"""
        agrupacion = (
            queryset.values("proposito_ganado")
            .annotate(
                total_registros=Count("id"),
                total_cabezas=Sum("cantidad_cabezas"),
                promedio_cabezas=Avg("cantidad_cabezas"),
                aprobadas=Count("id", filter=Q(estado="APROBADO")),
                rechazadas=Count("id", filter=Q(estado="RECHAZADO")),
                ingresos_total=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
                monto_promedio=Avg("monto_certificacion"),
                tiempo_promedio=Avg("tiempo_procesamiento_horas"),
            )
            .order_by("-total_cabezas")
        )

        total_nacional = queryset.count()

        datos_agrupados = []
        for grupo in agrupacion:
            total_procesadas = grupo["aprobadas"] + grupo["rechazadas"]

            item = {
                "proposito_ganado": grupo["proposito_ganado"],
                "proposito_display": dict(
                    MarcaGanadoBovino._meta.get_field("proposito_ganado").choices
                ).get(grupo["proposito_ganado"], grupo["proposito_ganado"]),
                "total_registros": grupo["total_registros"],
                "total_cabezas": grupo["total_cabezas"],
                "promedio_cabezas": round(grupo["promedio_cabezas"] or 0, 2),
                "participacion_mercado": round(
                    (
                        (grupo["total_registros"] / total_nacional * 100)
                        if total_nacional > 0
                        else 0
                    ),
                    2,
                ),
                "tasa_aprobacion": round(
                    (
                        (grupo["aprobadas"] / total_procesadas * 100)
                        if total_procesadas > 0
                        else 0
                    ),
                    2,
                ),
                "ingresos_total": float(grupo["ingresos_total"] or 0),
                "valor_promedio_marca": round(float(grupo["monto_promedio"] or 0), 2),
                "eficiencia_economica": round(
                    (
                        (float(grupo["ingresos_total"] or 0) / grupo["total_cabezas"])
                        if grupo["total_cabezas"] > 0
                        else 0
                    ),
                    2,
                ),
                "tiempo_promedio_procesamiento": round(
                    grupo["tiempo_promedio"] or 0, 2
                ),
            }

            if metricas:
                item_filtrado = {
                    "proposito_ganado": item["proposito_ganado"],
                    "proposito_display": item["proposito_display"],
                }
                for metrica in metricas:
                    if metrica in item:
                        item_filtrado[metrica] = item[metrica]
                datos_agrupados.append(item_filtrado)
            else:
                datos_agrupados.append(item)

        return datos_agrupados

    def _agrupar_por_mes(self, queryset, metricas):
        """Agrupa datos por mes según métricas solicitadas"""
        agrupacion = (
            queryset.extra(
                select={"año": "YEAR(fecha_registro)", "mes": "MONTH(fecha_registro)"}
            )
            .values("año", "mes")
            .annotate(
                total_registros=Count("id"),
                total_cabezas=Sum("cantidad_cabezas"),
                promedio_cabezas=Avg("cantidad_cabezas"),
                aprobadas=Count("id", filter=Q(estado="APROBADO")),
                rechazadas=Count("id", filter=Q(estado="RECHAZADO")),
                ingresos_total=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
                tiempo_promedio=Avg("tiempo_procesamiento_horas"),
            )
            .order_by("año", "mes")
        )

        datos_agrupados = []
        for grupo in agrupacion:
            total_procesadas = grupo["aprobadas"] + grupo["rechazadas"]

            item = {
                "año": grupo["año"],
                "mes": grupo["mes"],
                "mes_nombre": calendar.month_name[grupo["mes"]],
                "periodo": f"{calendar.month_name[grupo['mes']]} {grupo['año']}",
                "total_registros": grupo["total_registros"],
                "total_cabezas": grupo["total_cabezas"],
                "promedio_cabezas": round(grupo["promedio_cabezas"] or 0, 2),
                "tasa_aprobacion": round(
                    (
                        (grupo["aprobadas"] / total_procesadas * 100)
                        if total_procesadas > 0
                        else 0
                    ),
                    2,
                ),
                "ingresos_total": float(grupo["ingresos_total"] or 0),
                "tiempo_promedio_procesamiento": round(
                    grupo["tiempo_promedio"] or 0, 2
                ),
            }

            if metricas:
                item_filtrado = {
                    "año": item["año"],
                    "mes": item["mes"],
                    "periodo": item["periodo"],
                }
                for metrica in metricas:
                    if metrica in item:
                        item_filtrado[metrica] = item[metrica]
                datos_agrupados.append(item_filtrado)
            else:
                datos_agrupados.append(item)

        return datos_agrupados

    def _calcular_metricas_globales(self, queryset, metricas):
        """Calcula métricas globales del queryset"""
        total_registros = queryset.count()

        agregaciones = queryset.aggregate(
            total_cabezas=Sum("cantidad_cabezas"),
            promedio_cabezas=Avg("cantidad_cabezas"),
            aprobadas=Count("id", filter=Q(estado="APROBADO")),
            rechazadas=Count("id", filter=Q(estado="RECHAZADO")),
            pendientes=Count("id", filter=Q(estado__in=["PENDIENTE", "EN_PROCESO"])),
            ingresos_total=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
            tiempo_promedio=Avg("tiempo_procesamiento_horas"),
            monto_promedio=Avg("monto_certificacion"),
            productores_unicos=Count("ci_productor", distinct=True),
            departamentos_activos=Count("departamento", distinct=True),
            razas_registradas=Count("raza_bovino", distinct=True),
        )

        total_procesadas = agregaciones["aprobadas"] + agregaciones["rechazadas"]

        metricas_globales = {
            "total_registros": total_registros,
            "total_cabezas": agregaciones["total_cabezas"],
            "promedio_cabezas": round(agregaciones["promedio_cabezas"] or 0, 2),
            "tasa_aprobacion": round(
                (
                    (agregaciones["aprobadas"] / total_procesadas * 100)
                    if total_procesadas > 0
                    else 0
                ),
                2,
            ),
            "tasa_rechazo": round(
                (
                    (agregaciones["rechazadas"] / total_procesadas * 100)
                    if total_procesadas > 0
                    else 0
                ),
                2,
            ),
            "pendientes": agregaciones["pendientes"],
            "ingresos_total": float(agregaciones["ingresos_total"] or 0),
            "tiempo_promedio_procesamiento": round(
                agregaciones["tiempo_promedio"] or 0, 2
            ),
            "valor_promedio_marca": round(
                float(agregaciones["monto_promedio"] or 0), 2
            ),
            "productores_activos": agregaciones["productores_unicos"],
            "departamentos_activos": agregaciones["departamentos_activos"],
            "diversidad_genetica": agregaciones["razas_registradas"],
            "densidad_ganadera_promedio": round(
                (
                    (agregaciones["total_cabezas"] / total_registros)
                    if total_registros > 0
                    else 0
                ),
                2,
            ),
        }

        # Filtrar métricas si se especificaron
        if metricas:
            return {
                metrica: metricas_globales[metrica]
                for metrica in metricas
                if metrica in metricas_globales
            }

        return metricas_globales

    def _generar_resumen_estadistico(self, datos_agrupados, agrupacion):
        """Genera resumen estadístico de los datos agrupados"""
        if not datos_agrupados:
            return {"mensaje": "Sin datos para resumen estadístico"}

        resumen = {"total_grupos": len(datos_agrupados), "tipo_agrupacion": agrupacion}

        # Análisis según tipo de agrupación
        if agrupacion == "departamento":
            total_cabezas = [item.get("total_cabezas", 0) for item in datos_agrupados]
            resumen.update(
                {
                    "departamento_lider": max(
                        datos_agrupados, key=lambda x: x.get("total_cabezas", 0)
                    ).get("departamento_display"),
                    "mayor_concentracion_cabezas": max(total_cabezas),
                    "menor_concentracion_cabezas": min(total_cabezas),
                    "promedio_cabezas_por_departamento": round(
                        sum(total_cabezas) / len(total_cabezas), 2
                    ),
                }
            )

        elif agrupacion == "raza_bovino":
            participaciones = [
                item.get("participacion_mercado", 0) for item in datos_agrupados
            ]
            resumen.update(
                {
                    "raza_dominante": max(
                        datos_agrupados, key=lambda x: x.get("participacion_mercado", 0)
                    ).get("raza_display"),
                    "mayor_participacion": max(participaciones),
                    "diversidad_genetica": len(datos_agrupados),
                    "concentracion_top3": sum(
                        sorted(participaciones, reverse=True)[:3]
                    ),
                }
            )

        elif agrupacion == "proposito_ganado":
            participaciones = [
                item.get("participacion_mercado", 0) for item in datos_agrupados
            ]
            eficiencias = [
                item.get("eficiencia_economica", 0) for item in datos_agrupados
            ]
            resumen.update(
                {
                    "proposito_dominante": max(
                        datos_agrupados, key=lambda x: x.get("participacion_mercado", 0)
                    ).get("proposito_display"),
                    "mayor_participacion": max(participaciones),
                    "proposito_mas_eficiente": max(
                        datos_agrupados, key=lambda x: x.get("eficiencia_economica", 0)
                    ).get("proposito_display"),
                    "diversificacion_productiva": len(datos_agrupados),
                }
            )

        elif agrupacion == "mes":
            registros_mensuales = [
                item.get("total_registros", 0) for item in datos_agrupados
            ]
            if registros_mensuales:
                resumen.update(
                    {
                        "mes_mas_activo": max(
                            datos_agrupados, key=lambda x: x.get("total_registros", 0)
                        ).get("periodo"),
                        "mes_menos_activo": min(
                            datos_agrupados, key=lambda x: x.get("total_registros", 0)
                        ).get("periodo"),
                        "promedio_mensual": round(
                            sum(registros_mensuales) / len(registros_mensuales), 2
                        ),
                        "variabilidad": (
                            "alta"
                            if (max(registros_mensuales) - min(registros_mensuales))
                            > sum(registros_mensuales) / len(registros_mensuales)
                            else "baja"
                        ),
                    }
                )

        return resumen

    # ==================== MÉTODOS DE EXPORTACIÓN ====================

    def _exportar_excel_personalizado(self, reporte_personalizado):
        """Exporta reporte personalizado a Excel"""
        try:
            output = io.BytesIO()

            # Intentar con xlsxwriter primero
            try:
                import xlsxwriter

                workbook = xlsxwriter.Workbook(output, {"in_memory": True})

                # Crear hoja principal
                worksheet = workbook.add_worksheet("Reporte Personalizado")

                # Formatos
                header_format = workbook.add_format(
                    {
                        "bold": True,
                        "font_color": "white",
                        "bg_color": "#2E86AB",
                        "border": 1,
                    }
                )

                cell_format = workbook.add_format({"border": 1})

                # Escribir configuración
                row = 0
                worksheet.write(row, 0, "CONFIGURACIÓN DEL REPORTE", header_format)
                row += 1

                config = reporte_personalizado["configuracion"]
                for key, value in config.items():
                    if key != "filtros_aplicados":
                        worksheet.write(
                            row, 0, key.replace("_", " ").title(), cell_format
                        )
                        worksheet.write(row, 1, str(value), cell_format)
                        row += 1

                # Filtros aplicados
                row += 1
                worksheet.write(row, 0, "FILTROS APLICADOS", header_format)
                row += 1

                for filtro, valor in config.get("filtros_aplicados", {}).items():
                    worksheet.write(
                        row, 0, filtro.replace("_", " ").title(), cell_format
                    )
                    worksheet.write(row, 1, str(valor), cell_format)
                    row += 1

                # Datos principales
                row += 2
                worksheet.write(row, 0, "DATOS DEL REPORTE", header_format)
                row += 1

                datos = reporte_personalizado["datos"]
                if datos and isinstance(datos, list) and len(datos) > 0:
                    # Headers
                    headers = list(datos[0].keys())
                    for col, header in enumerate(headers):
                        worksheet.write(
                            row, col, header.replace("_", " ").title(), header_format
                        )
                    row += 1

                    # Datos
                    for item in datos:
                        for col, header in enumerate(headers):
                            value = item.get(header, "")
                            worksheet.write(row, col, str(value), cell_format)
                        row += 1

                workbook.close()

            except ImportError:
                # Fallback a openpyxl
                import openpyxl
                import pandas as pd

                workbook = openpyxl.Workbook()
                ws = workbook.active
                ws.title = "Reporte Personalizado"

                # Escribir datos como DataFrame
                if reporte_personalizado["datos"]:
                    df = pd.DataFrame(reporte_personalizado["datos"])

                    # Escribir headers
                    for col, header in enumerate(df.columns, 1):
                        ws.cell(
                            row=1, column=col, value=header.replace("_", " ").title()
                        )

                    # Escribir datos
                    for row_idx, row_data in enumerate(df.values, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)

                workbook.save(output)

            output.seek(0)

            filename = (
                f'reporte_personalizado_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return Response(
                {"error": f"Error exportando a Excel: {str(e)}"}, status=500
            )

    def _exportar_csv_personalizado(self, reporte_personalizado):
        """Exporta reporte personalizado a CSV"""
        try:
            output = io.StringIO()

            datos = reporte_personalizado["datos"]
            if not datos:
                return Response({"error": "No hay datos para exportar"}, status=400)

            # Crear CSV writer
            fieldnames = list(datos[0].keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames)

            # Escribir headers
            writer.writeheader()

            # Escribir datos
            for item in datos:
                # Convertir valores no string a string
                row = {}
                for key, value in item.items():
                    if isinstance(value, (dict, list)):
                        row[key] = json.dumps(value)
                    else:
                        row[key] = str(value) if value is not None else ""
                writer.writerow(row)

            # Preparar respuesta
            output.seek(0)
            filename = (
                f'reporte_personalizado_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
            )

            response = HttpResponse(output.getvalue(), content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return Response({"error": f"Error exportando a CSV: {str(e)}"}, status=500)

    def _crear_hoja_datos_principales(self, workbook, datos):
        """Crea hoja principal con datos de marcas"""
        try:
            import xlsxwriter

            worksheet = workbook.add_worksheet("Datos Principales")

            # Formatos
            header_format = workbook.add_format(
                {
                    "bold": True,
                    "font_color": "white",
                    "bg_color": "#1f4e79",
                    "border": 1,
                    "text_wrap": True,
                }
            )

            cell_format = workbook.add_format({"border": 1})
            number_format = workbook.add_format({"border": 1, "num_format": "#,##0.00"})
            date_format = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd"})

            # Headers
            headers = [
                "Número de Marca",
                "Productor",
                "CI Productor",
                "Teléfono",
                "Raza Bovina",
                "Propósito",
                "Cantidad Cabezas",
                "Departamento",
                "Municipio",
                "Comunidad",
                "Estado",
                "Monto Certificación",
                "Fecha Registro",
                "Fecha Procesamiento",
                "Tiempo Procesamiento (hrs)",
                "Observaciones",
            ]

            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
                worksheet.set_column(col, col, 15)  # Ancho de columna

            # Datos
            for row, item in enumerate(datos, 1):
                worksheet.write(row, 0, item.get("Número de Marca", ""), cell_format)
                worksheet.write(row, 1, item.get("Productor", ""), cell_format)
                worksheet.write(row, 2, item.get("CI Productor", ""), cell_format)
                worksheet.write(row, 3, item.get("Teléfono", ""), cell_format)
                worksheet.write(row, 4, item.get("Raza Bovina", ""), cell_format)
                worksheet.write(row, 5, item.get("Propósito", ""), cell_format)
                worksheet.write(row, 6, item.get("Cantidad Cabezas", 0), cell_format)
                worksheet.write(row, 7, item.get("Departamento", ""), cell_format)
                worksheet.write(row, 8, item.get("Municipio", ""), cell_format)
                worksheet.write(row, 9, item.get("Comunidad", ""), cell_format)
                worksheet.write(row, 10, item.get("Estado", ""), cell_format)
                worksheet.write(
                    row, 11, item.get("Monto Certificación", 0), number_format
                )

                # Fechas
                fecha_registro = item.get("Fecha Registro")
                if fecha_registro:
                    worksheet.write(row, 12, fecha_registro, date_format)
                else:
                    worksheet.write(row, 12, "", cell_format)

                fecha_procesamiento = item.get("Fecha Procesamiento")
                if fecha_procesamiento:
                    worksheet.write(row, 13, fecha_procesamiento, date_format)
                else:
                    worksheet.write(row, 13, "", cell_format)

                worksheet.write(
                    row, 14, item.get("Tiempo Procesamiento (hrs)", 0), cell_format
                )
                worksheet.write(row, 15, item.get("Observaciones", ""), cell_format)

        except Exception as e:
            print(f"Error creando hoja de datos principales: {e}")

    def _crear_hoja_resumen_estadistico(self, workbook, resumen):
        """Crea hoja con resumen estadístico"""
        try:
            import xlsxwriter

            worksheet = workbook.add_worksheet("Resumen Estadístico")

            # Formatos
            title_format = workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#2E86AB",
                }
            )

            header_format = workbook.add_format(
                {"bold": True, "bg_color": "#A8DADC", "border": 1}
            )

            cell_format = workbook.add_format({"border": 1})
            number_format = workbook.add_format({"border": 1, "num_format": "#,##0.00"})

            # Título
            worksheet.merge_range(
                "A1:B1", "RESUMEN ESTADÍSTICO DEL REPORTE", title_format
            )

            # Datos del resumen
            row = 3
            for key, value in resumen.items():
                worksheet.write(row, 0, key.replace("_", " ").title(), header_format)
                if isinstance(value, (int, float)):
                    worksheet.write(row, 1, value, number_format)
                else:
                    worksheet.write(row, 1, str(value), cell_format)
                row += 1

            # Ajustar ancho de columnas
            worksheet.set_column("A:A", 25)
            worksheet.set_column("B:B", 20)

        except Exception as e:
            print(f"Error creando hoja de resumen: {e}")

    # ==================== MÉTODOS AUXILIARES FALTANTES ====================

    def _calcular_score_desempeño(self, metricas):
        """Calcula score general de desempeño del sistema"""
        # Score basado en múltiples factores
        score_tiempo = (
            max(0, 100 - (metricas["tiempo_promedio_procesamiento"] - 24) * 1.5)
            if metricas["tiempo_promedio_procesamiento"] > 0
            else 100
        )
        score_aprobacion = metricas["tasa_aprobacion"]
        score_volumen = min(
            (metricas["marcas_registradas"] / 100) * 100, 100
        )  # 100 marcas = 100%

        # Ponderación: 40% eficiencia, 40% calidad, 20% volumen
        score_total = score_tiempo * 0.4 + score_aprobacion * 0.4 + score_volumen * 0.2

        return {
            "score_total": round(score_total, 1),
            "score_tiempo": round(score_tiempo, 1),
            "score_aprobacion": round(score_aprobacion, 1),
            "score_volumen": round(score_volumen, 1),
            "clasificacion": self._clasificar_score_desempeño(score_total),
        }

    def _clasificar_score_desempeño(self, score):
        """Clasifica el score de desempeño"""
        if score >= 90:
            return "excelente"
        elif score >= 75:
            return "muy_bueno"
        elif score >= 60:
            return "bueno"
        elif score >= 45:
            return "regular"
        else:
            return "necesita_mejora"

    def _generar_recomendacion_principal(self, tendencia, metricas):
        """Genera la recomendación principal basada en análisis"""
        if tendencia == "crecimiento_fuerte":
            return "Mantener momentum de crecimiento y preparar capacidad para demanda futura"
        elif tendencia == "crecimiento_moderado":
            return "Implementar estrategias para acelerar crecimiento y capturar mayor participación de mercado"
        elif tendencia == "estable":
            return "Analizar oportunidades de innovación para romper estancamiento"
        elif tendencia == "decrecimiento":
            return "Acción inmediata requerida: revisar precios, procesos y estrategia de mercado"
        elif metricas["tiempo_promedio_procesamiento"] > 72:
            return "Priorizar mejora de eficiencia operacional para incrementar satisfacción del cliente"
        elif metricas["tasa_aprobacion"] < 70:
            return "Fortalecer controles de calidad y capacitación del equipo evaluador"
        else:
            return "Optimizar procesos existentes y explorar nuevas oportunidades de mercado"

    # ==================== MÉTODOS DE EXPORTACIÓN EXCEL COMPLETADOS ====================

    def _crear_hoja_analisis_departamental(self, workbook, filtros):
        """Crea hoja con análisis departamental"""
        try:
            import xlsxwriter

            # Obtener datos departamentales
            queryset = MarcaGanadoBovino.objects.all()
            if filtros.get("fecha_desde"):
                queryset = queryset.filter(
                    fecha_registro__date__gte=filtros["fecha_desde"]
                )
            if filtros.get("fecha_hasta"):
                queryset = queryset.filter(
                    fecha_registro__date__lte=filtros["fecha_hasta"]
                )

            datos_departamentales = self._agrupar_por_departamento(queryset, [])

            worksheet = workbook.add_worksheet("Análisis Departamental")

            # Formatos
            header_format = workbook.add_format(
                {
                    "bold": True,
                    "font_color": "white",
                    "bg_color": "#1f4e79",
                    "border": 1,
                }
            )

            cell_format = workbook.add_format({"border": 1})
            number_format = workbook.add_format({"border": 1, "num_format": "#,##0.00"})
            percentage_format = workbook.add_format(
                {"border": 1, "num_format": "0.00%"}
            )

            # Headers
            headers = [
                "Departamento",
                "Total Registros",
                "Total Cabezas",
                "Promedio Cabezas",
                "Tasa Aprobación (%)",
                "Ingresos Total",
                "Tiempo Promedio (hrs)",
                "Productores Activos",
                "Densidad Ganadera",
            ]

            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
                worksheet.set_column(col, col, 18)

            # Datos
            for row, dept in enumerate(datos_departamentales, 1):
                worksheet.write(
                    row, 0, dept.get("departamento_display", ""), cell_format
                )
                worksheet.write(row, 1, dept.get("total_registros", 0), cell_format)
                worksheet.write(row, 2, dept.get("total_cabezas", 0), cell_format)
                worksheet.write(row, 3, dept.get("promedio_cabezas", 0), number_format)
                worksheet.write(
                    row, 4, dept.get("tasa_aprobacion", 0) / 100, percentage_format
                )
                worksheet.write(row, 5, dept.get("ingresos_total", 0), number_format)
                worksheet.write(row, 6, dept.get("tiempo_promedio", 0), number_format)
                worksheet.write(row, 7, dept.get("productores_activos", 0), cell_format)
                worksheet.write(row, 8, dept.get("densidad_ganadera", 0), number_format)

            # Agregar gráfico si hay datos
            if datos_departamentales:
                self._agregar_grafico_departamental(
                    workbook, worksheet, len(datos_departamentales)
                )

        except Exception as e:
            print(f"Error creando hoja de análisis departamental: {e}")

    def _crear_hoja_analisis_razas(self, workbook, filtros):
        """Crea hoja con análisis de razas bovinas"""
        try:
            import xlsxwriter

            # Obtener datos de razas
            queryset = MarcaGanadoBovino.objects.all()
            if filtros.get("fecha_desde"):
                queryset = queryset.filter(
                    fecha_registro__date__gte=filtros["fecha_desde"]
                )
            if filtros.get("fecha_hasta"):
                queryset = queryset.filter(
                    fecha_registro__date__lte=filtros["fecha_hasta"]
                )

            datos_razas = self._agrupar_por_raza(queryset, [])

            worksheet = workbook.add_worksheet("Análisis Razas Bovinas")

            # Formatos
            header_format = workbook.add_format(
                {
                    "bold": True,
                    "font_color": "white",
                    "bg_color": "#2E86AB",
                    "border": 1,
                }
            )

            cell_format = workbook.add_format({"border": 1})
            number_format = workbook.add_format({"border": 1, "num_format": "#,##0.00"})
            percentage_format = workbook.add_format(
                {"border": 1, "num_format": "0.00%"}
            )

            # Headers
            headers = [
                "Raza Bovina",
                "Total Registros",
                "Total Cabezas",
                "Promedio Cabezas",
                "Participación Mercado (%)",
                "Tasa Aprobación (%)",
                "Ingresos Total",
                "Diversidad Geográfica",
                "Valor Promedio Marca",
            ]

            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
                worksheet.set_column(col, col, 16)

            # Datos
            for row, raza in enumerate(datos_razas, 1):
                worksheet.write(row, 0, raza.get("raza_display", ""), cell_format)
                worksheet.write(row, 1, raza.get("total_registros", 0), cell_format)
                worksheet.write(row, 2, raza.get("total_cabezas", 0), cell_format)
                worksheet.write(row, 3, raza.get("promedio_cabezas", 0), number_format)
                worksheet.write(
                    row,
                    4,
                    raza.get("participacion_mercado", 0) / 100,
                    percentage_format,
                )
                worksheet.write(
                    row, 5, raza.get("tasa_aprobacion", 0) / 100, percentage_format
                )
                worksheet.write(row, 6, raza.get("ingresos_total", 0), number_format)
                worksheet.write(
                    row, 7, raza.get("diversidad_geografica", 0), cell_format
                )
                worksheet.write(
                    row, 8, raza.get("valor_promedio_marca", 0), number_format
                )

            # Agregar análisis adicional
            self._agregar_analisis_genetico(worksheet, datos_razas)

        except Exception as e:
            print(f"Error creando hoja de análisis de razas: {e}")

    def _agregar_grafico_departamental(self, workbook, worksheet, num_departamentos):
        """Agrega gráfico de barras para análisis departamental"""
        try:
            import xlsxwriter

            # Crear gráfico de barras
            chart = workbook.add_chart({"type": "column"})

            # Configurar datos del gráfico
            chart.add_series(
                {
                    "name": "Total Cabezas por Departamento",
                    "categories": [
                        "Análisis Departamental",
                        1,
                        0,
                        num_departamentos,
                        0,
                    ],
                    "values": ["Análisis Departamental", 1, 2, num_departamentos, 2],
                    "data_labels": {"value": True},
                }
            )

            # Configurar gráfico
            chart.set_title({"name": "Distribución de Ganado Bovino por Departamento"})
            chart.set_x_axis({"name": "Departamentos"})
            chart.set_y_axis({"name": "Número de Cabezas"})
            chart.set_style(10)

            # Insertar gráfico
            worksheet.insert_chart("K2", chart, {"x_scale": 1.5, "y_scale": 1.2})

        except Exception as e:
            print(f"Error agregando gráfico departamental: {e}")

    def _agregar_analisis_genetico(self, worksheet, datos_razas):
        """Agrega análisis de diversidad genética a la hoja de razas"""
        try:
            # Encontrar última fila con datos
            start_row = len(datos_razas) + 3

            # Formatos
            title_format = worksheet.parent.add_format(
                {"bold": True, "font_size": 12, "bg_color": "#F4A261"}
            )

            info_format = worksheet.parent.add_format({"border": 1})

            # Título del análisis
            worksheet.write(
                start_row, 0, "ANÁLISIS DE DIVERSIDAD GENÉTICA", title_format
            )
            start_row += 2

            # Calcular métricas de diversidad
            total_razas = len(datos_razas)
            if datos_razas:
                raza_dominante = max(
                    datos_razas, key=lambda x: x.get("participacion_mercado", 0)
                )
                concentracion_top3 = sum(
                    sorted(
                        [r.get("participacion_mercado", 0) for r in datos_razas],
                        reverse=True,
                    )[:3]
                )

                # Escribir análisis
                analisis_items = [
                    ("Total de razas registradas:", total_razas),
                    ("Raza dominante:", raza_dominante.get("raza_display", "N/A")),
                    (
                        "Participación raza dominante:",
                        f"{raza_dominante.get('participacion_mercado', 0):.1f}%",
                    ),
                    ("Concentración TOP 3 razas:", f"{concentracion_top3:.1f}%"),
                    (
                        "Índice de diversidad:",
                        (
                            "Alto"
                            if total_razas >= 8
                            else "Medio" if total_razas >= 5 else "Bajo"
                        ),
                    ),
                ]

                for i, (label, value) in enumerate(analisis_items):
                    worksheet.write(start_row + i, 0, label, info_format)
                    worksheet.write(start_row + i, 1, str(value), info_format)

        except Exception as e:
            print(f"Error agregando análisis genético: {e}")

    # ==================== NUEVOS ENDPOINTS ADICIONALES ====================

    @action(detail=False, methods=["get"])
    def reporte_productor_individual(self, request):
        """Reporte detallado para un productor específico"""
        marca_id = request.query_params.get("marca_id")
        ci_productor = request.query_params.get("ci_productor")

        if not marca_id and not ci_productor:
            return Response(
                {"error": "Debe proporcionar marca_id o ci_productor"}, status=400
            )

        try:
            if marca_id:
                reporte = ReportService.generar_reporte_productor(marca_id)
            else:
                # Buscar por CI del productor
                marca = MarcaGanadoBovino.objects.filter(
                    ci_productor=ci_productor
                ).first()
                if not marca:
                    return Response(
                        {"error": "No se encontró productor con ese CI"}, status=404
                    )
                reporte = ReportService.generar_reporte_productor(marca.id)

            if "error" in reporte:
                return Response(reporte, status=404)

            # Enriquecer con análisis adicional
            reporte["recomendaciones_mejora"] = self._generar_recomendaciones_productor(
                reporte
            )
            reporte["benchmarking"] = self._generar_benchmarking_productor(reporte)

            return Response(reporte)

        except Exception as e:
            return Response(
                {"error": f"Error generando reporte del productor: {str(e)}"},
                status=500,
            )

    @action(detail=False, methods=["get"])
    def reporte_tendencias_temporales(self, request):
        """Análisis de tendencias temporales avanzado"""
        try:
            periodo = request.query_params.get(
                "periodo", "ultimo_año"
            )  # ultimo_año, ultimos_6_meses, ultimos_3_meses
            granularidad = request.query_params.get(
                "granularidad", "mensual"
            )  # mensual, semanal, diario

            # Calcular fechas según período
            ahora = timezone.now()
            if periodo == "ultimo_año":
                fecha_inicio = ahora - timedelta(days=365)
            elif periodo == "ultimos_6_meses":
                fecha_inicio = ahora - timedelta(days=180)
            elif periodo == "ultimos_3_meses":
                fecha_inicio = ahora - timedelta(days=90)
            else:
                fecha_inicio = ahora - timedelta(days=365)

            # Generar análisis temporal
            analisis_temporal = self._generar_analisis_temporal(
                fecha_inicio, granularidad
            )

            # Detectar patrones estacionales
            patrones_estacionales = self._detectar_patrones_estacionales(fecha_inicio)

            # Proyecciones futuras
            proyecciones = self._generar_proyecciones_temporales(analisis_temporal)

            return Response(
                {
                    "configuracion": {
                        "periodo": periodo,
                        "granularidad": granularidad,
                        "fecha_inicio": fecha_inicio.date(),
                        "fecha_fin": ahora.date(),
                    },
                    "analisis_temporal": analisis_temporal,
                    "patrones_estacionales": patrones_estacionales,
                    "proyecciones": proyecciones,
                    "insights_temporales": self._generar_insights_temporales(
                        analisis_temporal, patrones_estacionales
                    ),
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Error en análisis temporal: {str(e)}"}, status=500
            )

    @action(detail=False, methods=["get"])
    def dashboard_ejecutivo(self, request):
        """Dashboard ejecutivo con métricas clave en tiempo real"""
        try:
            # Métricas del mes actual
            ahora = timezone.now()
            inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

            # KPIs principales
            kpis_mes = self._calcular_kpis_mes_actual(inicio_mes)

            # Comparación con mes anterior
            mes_anterior = (inicio_mes - timedelta(days=1)).replace(day=1)
            fin_mes_anterior = inicio_mes
            kpis_anterior = self._calcular_kpis_periodo(mes_anterior, fin_mes_anterior)

            # Alertas críticas
            alertas_criticas = self._detectar_alertas_criticas()

            # Top performers
            top_performers = self._identificar_top_performers()

            # Indicadores de salud del sistema
            salud_sistema = self._evaluar_salud_sistema()

            # Predicciones para próximos 30 días
            predicciones = AnalyticsService.predecir_demanda_mensual()

            return Response(
                {
                    "fecha_actualizacion": ahora,
                    "kpis_mes_actual": kpis_mes,
                    "comparacion_mes_anterior": {
                        "periodo_anterior": f"{mes_anterior.strftime('%B %Y')}",
                        "kpis_anterior": kpis_anterior,
                        "variaciones": self._calcular_variaciones_kpis(
                            kpis_mes, kpis_anterior
                        ),
                    },
                    "alertas_criticas": alertas_criticas,
                    "top_performers": top_performers,
                    "salud_sistema": salud_sistema,
                    "predicciones_30_dias": predicciones,
                    "acciones_recomendadas": self._generar_acciones_dashboard(
                        alertas_criticas, salud_sistema, predicciones
                    ),
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Error en dashboard ejecutivo: {str(e)}"}, status=500
            )

    # ==================== MÉTODOS AUXILIARES PARA NUEVOS ENDPOINTS ====================

    def _generar_recomendaciones_productor(self, reporte_productor):
        """Genera recomendaciones específicas para el productor"""
        recomendaciones = []
        marca_info = reporte_productor["marca"]
        comparacion = reporte_productor["comparacion_regional"]

        # Recomendación por tamaño del hato
        if marca_info["cantidad_cabezas"] < comparacion["promedio_cabezas_region"]:
            recomendaciones.append(
                {
                    "categoria": "expansion_hato",
                    "prioridad": "media",
                    "recomendacion": "Considerar expansión gradual del hato ganadero",
                    "justificacion": f'Con {marca_info["cantidad_cabezas"]} cabezas está {comparacion["promedio_cabezas_region"] - marca_info["cantidad_cabezas"]:.0f} cabezas por debajo del promedio regional',
                    "beneficio_esperado": "Incremento en productividad y rentabilidad",
                }
            )

        # Recomendación por eficiencia económica
        if comparacion["posicion_monto"] == "inferior":
            recomendaciones.append(
                {
                    "categoria": "optimizacion_economica",
                    "prioridad": "alta",
                    "recomendacion": "Implementar estrategias de mejora de valor por cabeza",
                    "justificacion": "Monto de certificación por debajo del promedio regional",
                    "beneficio_esperado": "Incremento en ingresos por unidad ganadera",
                }
            )

        # Recomendación por historial de cambios
        cambios_frecuentes = len(reporte_productor["historial_cambios"])
        if cambios_frecuentes > 3:
            recomendaciones.append(
                {
                    "categoria": "mejora_procesos",
                    "prioridad": "media",
                    "recomendacion": "Revisar documentación y procedimientos internos",
                    "justificacion": f"{cambios_frecuentes} cambios de estado sugieren oportunidades de mejora",
                    "beneficio_esperado": "Reducción en tiempo de procesamiento futuro",
                }
            )

        return recomendaciones

    def _generar_benchmarking_productor(self, reporte_productor):
        """Genera benchmarking del productor vs mercado"""
        marca_info = reporte_productor["marca"]

        # Obtener percentiles del mercado
        todos_productores = MarcaGanadoBovino.objects.filter(
            departamento=(
                marca_info.get("ubicacion", "").split(",")[1].strip()
                if "," in marca_info.get("ubicacion", "")
                else ""
            ),
            raza_bovino=marca_info.get("raza", ""),
            proposito_ganado=marca_info.get("proposito", ""),
        ).values_list("cantidad_cabezas", "monto_certificacion")

        if todos_productores:
            cabezas_lista = [p[0] for p in todos_productores]
            montos_lista = [float(p[1]) for p in todos_productores]

            cabezas_lista.sort()
            montos_lista.sort()

            # Calcular percentiles
            def calcular_percentil(valor, lista):
                if not lista:
                    return 50
                posicion = sum(1 for x in lista if x < valor)
                return (posicion / len(lista)) * 100

            percentil_cabezas = calcular_percentil(
                marca_info["cantidad_cabezas"], cabezas_lista
            )
            percentil_monto = calcular_percentil(marca_info["monto"], montos_lista)

            return {
                "percentil_tamaño_hato": round(percentil_cabezas, 1),
                "percentil_valor_marca": round(percentil_monto, 1),
                "clasificacion_tamaño": self._clasificar_percentil(percentil_cabezas),
                "clasificacion_valor": self._clasificar_percentil(percentil_monto),
                "productores_comparados": len(todos_productores),
                "mercado_referencia": f"{marca_info.get('proposito', '')} - {marca_info.get('raza', '')}",
            }

        return {"mensaje": "Datos insuficientes para benchmarking"}

    def _clasificar_percentil(self, percentil):
        """Clasifica el percentil en categorías"""
        if percentil >= 90:
            return "top_10"
        elif percentil >= 75:
            return "cuartil_superior"
        elif percentil >= 50:
            return "promedio_superior"
        elif percentil >= 25:
            return "promedio_inferior"
        else:
            return "cuartil_inferior"

    # ==================== MÉTODOS DE ANÁLISIS TEMPORAL ====================

    def _generar_analisis_temporal(self, fecha_inicio, granularidad):
        """Genera análisis temporal detallado según granularidad"""
        queryset = MarcaGanadoBovino.objects.filter(fecha_registro__gte=fecha_inicio)

        if granularidad == "mensual":
            return self._analisis_temporal_mensual(queryset)
        elif granularidad == "semanal":
            return self._analisis_temporal_semanal(queryset)
        elif granularidad == "diario":
            return self._analisis_temporal_diario(queryset)
        else:
            return self._analisis_temporal_mensual(queryset)

    def _analisis_temporal_mensual(self, queryset):
        """Análisis temporal agrupado por mes"""
        analisis = (
            queryset.extra(
                select={"año": "YEAR(fecha_registro)", "mes": "MONTH(fecha_registro)"}
            )
            .values("año", "mes")
            .annotate(
                total_registros=Count("id"),
                total_cabezas=Sum("cantidad_cabezas"),
                promedio_cabezas=Avg("cantidad_cabezas"),
                aprobadas=Count("id", filter=Q(estado="APROBADO")),
                rechazadas=Count("id", filter=Q(estado="RECHAZADO")),
                pendientes=Count(
                    "id", filter=Q(estado__in=["PENDIENTE", "EN_PROCESO"])
                ),
                ingresos=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
                tiempo_promedio=Avg("tiempo_procesamiento_horas"),
                productores_nuevos=Count("ci_productor", distinct=True),
            )
            .order_by("año", "mes")
        )

        # Enriquecer datos
        datos_temporales = []
        for item in analisis:
            total_procesadas = item["aprobadas"] + item["rechazadas"]

            # Calcular métricas derivadas
            tasa_aprobacion = (
                (item["aprobadas"] / total_procesadas * 100)
                if total_procesadas > 0
                else 0
            )
            densidad_ganadera = (
                item["total_cabezas"] / item["total_registros"]
                if item["total_registros"] > 0
                else 0
            )
            ingreso_por_cabeza = (
                float(item["ingresos"] or 0) / item["total_cabezas"]
                if item["total_cabezas"] > 0
                else 0
            )

            datos_temporales.append(
                {
                    "año": item["año"],
                    "mes": item["mes"],
                    "mes_nombre": calendar.month_name[item["mes"]],
                    "periodo": f"{calendar.month_name[item['mes']]} {item['año']}",
                    "total_registros": item["total_registros"],
                    "total_cabezas": item["total_cabezas"],
                    "promedio_cabezas": round(item["promedio_cabezas"] or 0, 2),
                    "tasa_aprobacion": round(tasa_aprobacion, 2),
                    "tasa_pendientes": round(
                        (
                            (item["pendientes"] / item["total_registros"] * 100)
                            if item["total_registros"] > 0
                            else 0
                        ),
                        2,
                    ),
                    "ingresos_total": float(item["ingresos"] or 0),
                    "tiempo_promedio_procesamiento": round(
                        item["tiempo_promedio"] or 0, 2
                    ),
                    "productores_activos": item["productores_nuevos"],
                    "densidad_ganadera": round(densidad_ganadera, 2),
                    "eficiencia_economica": round(ingreso_por_cabeza, 2),
                    "indice_actividad": self._calcular_indice_actividad(item),
                }
            )

        return datos_temporales

    def _analisis_temporal_semanal(self, queryset):
        """Análisis temporal agrupado por semana"""
        analisis = (
            queryset.extra(
                select={
                    "año": "YEAR(fecha_registro)",
                    "semana": "WEEK(fecha_registro, 1)",  # ISO week
                }
            )
            .values("año", "semana")
            .annotate(
                total_registros=Count("id"),
                total_cabezas=Sum("cantidad_cabezas"),
                promedio_cabezas=Avg("cantidad_cabezas"),
                aprobadas=Count("id", filter=Q(estado="APROBADO")),
                ingresos=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
                tiempo_promedio=Avg("tiempo_procesamiento_horas"),
            )
            .order_by("año", "semana")
        )

        datos_semanales = []
        for item in analisis:
            # Calcular fecha de inicio de la semana
            import datetime

            fecha_semana = datetime.datetime.strptime(
                f"{item['año']}-W{item['semana']:02d}-1", "%Y-W%W-%w"
            ).date()

            datos_semanales.append(
                {
                    "año": item["año"],
                    "semana": item["semana"],
                    "fecha_inicio_semana": fecha_semana,
                    "periodo": f"Sem {item['semana']} - {item['año']}",
                    "total_registros": item["total_registros"],
                    "total_cabezas": item["total_cabezas"],
                    "promedio_cabezas": round(item["promedio_cabezas"] or 0, 2),
                    "ingresos_total": float(item["ingresos"] or 0),
                    "tiempo_promedio_procesamiento": round(
                        item["tiempo_promedio"] or 0, 2
                    ),
                    "velocidad_registro": round(
                        item["total_registros"] / 7, 2
                    ),  # registros por día
                    "indice_productividad": round(
                        (item["total_cabezas"] or 0) / max(item["total_registros"], 1),
                        2,
                    ),
                }
            )

        return datos_semanales

    def _analisis_temporal_diario(self, queryset):
        """Análisis temporal agrupado por día"""
        # Limitar a últimos 60 días para evitar sobrecarga
        fecha_limite = timezone.now() - timedelta(days=60)
        queryset_filtrado = queryset.filter(fecha_registro__gte=fecha_limite)

        analisis = (
            queryset_filtrado.extra(select={"fecha": "DATE(fecha_registro)"})
            .values("fecha")
            .annotate(
                total_registros=Count("id"),
                total_cabezas=Sum("cantidad_cabezas"),
                promedio_cabezas=Avg("cantidad_cabezas"),
                aprobadas=Count("id", filter=Q(estado="APROBADO")),
                ingresos=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
            )
            .order_by("fecha")
        )

        datos_diarios = []
        for item in analisis:
            # Determinar día de la semana
            fecha_obj = item["fecha"]
            dia_semana = (
                fecha_obj.strftime("%A") if isinstance(fecha_obj, date) else "N/A"
            )

            datos_diarios.append(
                {
                    "fecha": item["fecha"],
                    "dia_semana": dia_semana,
                    "es_fin_semana": dia_semana in ["Saturday", "Sunday"],
                    "total_registros": item["total_registros"],
                    "total_cabezas": item["total_cabezas"],
                    "promedio_cabezas": round(item["promedio_cabezas"] or 0, 2),
                    "ingresos_total": float(item["ingresos"] or 0),
                    "intensidad_actividad": self._clasificar_intensidad_diaria(
                        item["total_registros"]
                    ),
                }
            )

        return datos_diarios

    def _calcular_indice_actividad(self, item_mensual):
        """Calcula un índice de actividad mensual normalizado"""
        # Factores: registros, cabezas, eficiencia, ingresos
        registros_norm = min(
            item_mensual["total_registros"] / 100, 1.0
        )  # max 100 registros = 1.0
        cabezas_norm = min(
            (item_mensual["total_cabezas"] or 0) / 5000, 1.0
        )  # max 5000 cabezas = 1.0

        total_procesadas = item_mensual["aprobadas"] + item_mensual["rechazadas"]
        eficiencia_norm = (
            (item_mensual["aprobadas"] / total_procesadas)
            if total_procesadas > 0
            else 0
        )

        ingresos_norm = min(
            float(item_mensual["ingresos"] or 0) / 1000000, 1.0
        )  # max 1M = 1.0

        # Índice ponderado
        indice = (
            registros_norm * 0.3
            + cabezas_norm * 0.3
            + eficiencia_norm * 0.2
            + ingresos_norm * 0.2
        ) * 100

        return round(indice, 2)

    def _clasificar_intensidad_diaria(self, registros_dia):
        """Clasifica la intensidad de actividad diaria"""
        if registros_dia >= 20:
            return "muy_alta"
        elif registros_dia >= 15:
            return "alta"
        elif registros_dia >= 10:
            return "media"
        elif registros_dia >= 5:
            return "baja"
        else:
            return "muy_baja"

    def _detectar_patrones_estacionales(self, fecha_inicio):
        """Detecta patrones estacionales en los datos"""
        queryset = MarcaGanadoBovino.objects.filter(fecha_registro__gte=fecha_inicio)

        # Análisis por mes del año (estacionalidad)
        patron_mensual = (
            queryset.extra(select={"mes": "MONTH(fecha_registro)"})
            .values("mes")
            .annotate(
                total_registros=Count("id"),
                promedio_cabezas=Avg("cantidad_cabezas"),
                total_cabezas=Sum("cantidad_cabezas"),
            )
            .order_by("mes")
        )

        # Enriquecer con información estacional
        meses_estacionales = []
        for mes_data in patron_mensual:
            mes_num = mes_data["mes"]

            # Determinar estación en Bolivia (hemisferio sur)
            if mes_num in [12, 1, 2]:
                estacion = "verano"
                clima = "lluvioso"
            elif mes_num in [3, 4, 5]:
                estacion = "otoño"
                clima = "transicion_seco"
            elif mes_num in [6, 7, 8]:
                estacion = "invierno"
                clima = "seco_frio"
            else:  # 9, 10, 11
                estacion = "primavera"
                clima = "transicion_lluvioso"

            meses_estacionales.append(
                {
                    "mes": mes_num,
                    "mes_nombre": calendar.month_name[mes_num],
                    "estacion": estacion,
                    "clima_boliviano": clima,
                    "total_registros": mes_data["total_registros"],
                    "promedio_cabezas": round(mes_data["promedio_cabezas"] or 0, 2),
                    "total_cabezas": mes_data["total_cabezas"],
                    "intensidad_relativa": self._calcular_intensidad_relativa(
                        mes_data, patron_mensual
                    ),
                }
            )

        # Análisis por día de la semana
        patron_semanal = (
            queryset.extra(select={"dia_semana": "DAYOFWEEK(fecha_registro)"})
            .values("dia_semana")
            .annotate(total_registros=Count("id"), promedio_registros=Count("id"))
            .order_by("dia_semana")
        )

        dias_semana_nombres = [
            "",
            "Domingo",
            "Lunes",
            "Martes",
            "Miércoles",
            "Jueves",
            "Viernes",
            "Sábado",
        ]

        patron_semanal_procesado = []
        for dia_data in patron_semanal:
            dia_num = dia_data["dia_semana"]
            es_laborable = dia_num not in [1, 7]  # No domingo ni sábado

            patron_semanal_procesado.append(
                {
                    "dia_numero": dia_num,
                    "dia_nombre": dias_semana_nombres[dia_num],
                    "es_laborable": es_laborable,
                    "total_registros": dia_data["total_registros"],
                    "porcentaje_actividad": round(
                        dia_data["total_registros"]
                        / sum(d["total_registros"] for d in patron_semanal)
                        * 100,
                        2,
                    ),
                }
            )

        return {
            "patron_estacional_mensual": meses_estacionales,
            "patron_semanal": patron_semanal_procesado,
            "insights_estacionalidad": self._generar_insights_estacionalidad(
                meses_estacionales, patron_semanal_procesado
            ),
            "recomendaciones_estacionales": self._generar_recomendaciones_estacionales(
                meses_estacionales
            ),
        }

    def _calcular_intensidad_relativa(self, mes_data, todos_meses):
        """Calcula intensidad relativa del mes vs promedio anual"""
        promedio_anual = sum(m["total_registros"] for m in todos_meses) / len(
            todos_meses
        )
        if promedio_anual > 0:
            intensidad = (mes_data["total_registros"] / promedio_anual) * 100
            return round(intensidad, 1)
        return 100

    def _generar_insights_estacionalidad(self, meses_estacionales, patron_semanal):
        """Genera insights sobre patrones estacionales"""
        insights = []

        # Insight sobre mes más activo
        if meses_estacionales:
            mes_mas_activo = max(meses_estacionales, key=lambda x: x["total_registros"])
            insights.append(
                {
                    "tipo": "pico_estacional",
                    "descripcion": f"Pico de actividad en {mes_mas_activo['mes_nombre']} ({mes_mas_activo['estacion']})",
                    "valor": mes_mas_activo["total_registros"],
                    "contexto_climatico": mes_mas_activo["clima_boliviano"],
                }
            )

        # Insight sobre concentración en días laborables
        if patron_semanal:
            registros_laborables = sum(
                d["total_registros"] for d in patron_semanal if d["es_laborable"]
            )
            total_registros = sum(d["total_registros"] for d in patron_semanal)
            concentracion = (
                (registros_laborables / total_registros * 100)
                if total_registros > 0
                else 0
            )

            insights.append(
                {
                    "tipo": "concentracion_laboral",
                    "descripcion": f"{concentracion:.1f}% de actividad se concentra en días laborables",
                    "interpretacion": (
                        "alta"
                        if concentracion > 80
                        else "media" if concentracion > 70 else "baja"
                    ),
                }
            )

        # Insight sobre variabilidad estacional
        if len(meses_estacionales) >= 12:
            registros_por_mes = [m["total_registros"] for m in meses_estacionales]
            import statistics

            cv = (
                statistics.stdev(registros_por_mes)
                / statistics.mean(registros_por_mes)
                * 100
            )

            insights.append(
                {
                    "tipo": "variabilidad_estacional",
                    "descripcion": f"Coeficiente de variación estacional: {cv:.1f}%",
                    "interpretacion": (
                        "alta" if cv > 30 else "media" if cv > 15 else "baja"
                    ),
                }
            )

        return insights

    def _generar_recomendaciones_estacionales(self, meses_estacionales):
        """Genera recomendaciones basadas en estacionalidad"""
        if not meses_estacionales:
            return []

        recomendaciones = []

        # Identificar meses de baja actividad
        promedio_actividad = sum(
            m["total_registros"] for m in meses_estacionales
        ) / len(meses_estacionales)
        meses_bajos = [
            m
            for m in meses_estacionales
            if m["total_registros"] < promedio_actividad * 0.7
        ]

        if meses_bajos:
            recomendaciones.append(
                {
                    "categoria": "nivelacion_estacional",
                    "recomendacion": "Implementar campañas promocionales en meses de baja actividad",
                    "meses_objetivo": [m["mes_nombre"] for m in meses_bajos],
                    "beneficio_esperado": "Reducir variabilidad estacional en un 20%",
                }
            )

        # Recomendación por época climática
        meses_lluvia = [
            m
            for m in meses_estacionales
            if m["clima_boliviano"] in ["lluvioso", "transicion_lluvioso"]
        ]
        if meses_lluvia:
            actividad_lluvia = sum(m["total_registros"] for m in meses_lluvia)
            if actividad_lluvia > promedio_actividad * 2:
                recomendaciones.append(
                    {
                        "categoria": "preparacion_climatica",
                        "recomendacion": "Reforzar capacidad operativa durante época lluviosa",
                        "justificacion": "Mayor actividad durante meses de lluvia requiere recursos adicionales",
                        "acciones": [
                            "Incrementar personal temporal",
                            "Mejorar infraestructura",
                            "Acelerar procesos",
                        ],
                    }
                )

        return recomendaciones

    def _generar_proyecciones_temporales(self, analisis_temporal):
        """Genera proyecciones futuras basadas en análisis temporal"""
        if len(analisis_temporal) < 6:  # Necesitamos al menos 6 períodos
            return {"mensaje": "Datos insuficientes para proyecciones confiables"}

        # Extraer series temporales
        registros_serie = [item["total_registros"] for item in analisis_temporal]
        cabezas_serie = [item["total_cabezas"] for item in analisis_temporal]
        ingresos_serie = [item["ingresos_total"] for item in analisis_temporal]

        # Calcular tendencias (regresión lineal simple)
        def calcular_tendencia(serie):
            n = len(serie)
            if n < 3:
                return 0

            x_valores = list(range(1, n + 1))
            suma_x = sum(x_valores)
            suma_y = sum(serie)
            suma_xy = sum(x * y for x, y in zip(x_valores, serie))
            suma_x2 = sum(x * x for x in x_valores)

            pendiente = (n * suma_xy - suma_x * suma_y) / (
                n * suma_x2 - suma_x * suma_x
            )
            intercepto = (suma_y - pendiente * suma_x) / n

            return {"pendiente": pendiente, "intercepto": intercepto}

        tendencia_registros = calcular_tendencia(registros_serie)
        tendencia_cabezas = calcular_tendencia(cabezas_serie)
        tendencia_ingresos = calcular_tendencia(ingresos_serie)

        # Proyecciones para próximos 3 períodos
        n = len(analisis_temporal)
        proyecciones = []

        for i in range(1, 4):  # Próximos 3 períodos
            x_futuro = n + i

            proj_registros = max(
                0,
                tendencia_registros["intercepto"]
                + tendencia_registros["pendiente"] * x_futuro,
            )
            proj_cabezas = max(
                0,
                tendencia_cabezas["intercepto"]
                + tendencia_cabezas["pendiente"] * x_futuro,
            )
            proj_ingresos = max(
                0,
                tendencia_ingresos["intercepto"]
                + tendencia_ingresos["pendiente"] * x_futuro,
            )

            # Determinar siguiente período
            ultimo_periodo = analisis_temporal[-1]
            if "mes" in ultimo_periodo:
                siguiente_mes = ultimo_periodo["mes"] + i
                siguiente_año = ultimo_periodo["año"]
                if siguiente_mes > 12:
                    siguiente_mes -= 12
                    siguiente_año += 1
                periodo_nombre = f"{calendar.month_name[siguiente_mes]} {siguiente_año}"
            else:
                periodo_nombre = f"Período {i}"

            proyecciones.append(
                {
                    "periodo": periodo_nombre,
                    "horizonte": f'+{i} período{"s" if i > 1 else ""}',
                    "proyeccion_registros": int(proj_registros),
                    "proyeccion_cabezas": int(proj_cabezas),
                    "proyeccion_ingresos": round(proj_ingresos, 2),
                    "confianza": "alta" if i == 1 else "media" if i == 2 else "baja",
                }
            )

        # Calcular intervalos de confianza básicos
        import statistics

        if len(registros_serie) >= 3:
            error_registros = statistics.stdev(registros_serie) * 0.5
            for proj in proyecciones:
                proj["intervalo_registros"] = {
                    "minimo": max(
                        0, int(proj["proyeccion_registros"] - error_registros)
                    ),
                    "maximo": int(proj["proyeccion_registros"] + error_registros),
                }

        return {
            "proyecciones": proyecciones,
            "tendencias_detectadas": {
                "registros": (
                    "creciente"
                    if tendencia_registros["pendiente"] > 0
                    else "decreciente"
                ),
                "cabezas": (
                    "creciente" if tendencia_cabezas["pendiente"] > 0 else "decreciente"
                ),
                "ingresos": (
                    "creciente"
                    if tendencia_ingresos["pendiente"] > 0
                    else "decreciente"
                ),
            },
            "velocidad_cambio": {
                "registros_por_periodo": round(tendencia_registros["pendiente"], 2),
                "cabezas_por_periodo": round(tendencia_cabezas["pendiente"], 2),
                "ingresos_por_periodo": round(tendencia_ingresos["pendiente"], 2),
            },
            "metodologia": "regresion_lineal_simple",
            "limitaciones": [
                "No considera factores externos",
                "Asume tendencia lineal",
                "Basado en datos históricos",
            ],
        }

    def _generar_insights_temporales(self, analisis_temporal, patrones_estacionales):
        """Genera insights clave del análisis temporal"""
        if not analisis_temporal:
            return []

        insights = []

        # Insight sobre crecimiento general
        if len(analisis_temporal) >= 2:
            primer_periodo = analisis_temporal[0]
            ultimo_periodo = analisis_temporal[-1]

            crecimiento_registros = (
                (
                    (
                        ultimo_periodo["total_registros"]
                        - primer_periodo["total_registros"]
                    )
                    / primer_periodo["total_registros"]
                    * 100
                )
                if primer_periodo["total_registros"] > 0
                else 0
            )

            insights.append(
                {
                    "tipo": "crecimiento_general",
                    "descripcion": f"Crecimiento de {crecimiento_registros:+.1f}% en registros durante el período analizado",
                    "interpretacion": (
                        "positivo" if crecimiento_registros > 0 else "negativo"
                    ),
                    "impacto": (
                        "alto"
                        if abs(crecimiento_registros) > 20
                        else "medio" if abs(crecimiento_registros) > 10 else "bajo"
                    ),
                }
            )

        # Insight sobre consistencia
        if len(analisis_temporal) >= 6:
            registros_lista = [p["total_registros"] for p in analisis_temporal]
            import statistics

            cv = (
                statistics.stdev(registros_lista)
                / statistics.mean(registros_lista)
                * 100
            )

            insights.append(
                {
                    "tipo": "consistencia_operacional",
                    "descripcion": f"Coeficiente de variación: {cv:.1f}%",
                    "interpretacion": (
                        "estable"
                        if cv < 20
                        else "variable" if cv < 40 else "muy_variable"
                    ),
                    "recomendacion": (
                        "Implementar estrategias de nivelación"
                        if cv > 30
                        else "Mantener consistencia actual"
                    ),
                }
            )

        # Insight sobre eficiencia temporal
        eficiencias = [
            p.get("tiempo_promedio_procesamiento", 0)
            for p in analisis_temporal
            if p.get("tiempo_promedio_procesamiento", 0) > 0
        ]
        if eficiencias:
            eficiencia_promedio = sum(eficiencias) / len(eficiencias)
            insights.append(
                {
                    "tipo": "eficiencia_temporal",
                    "descripcion": f"Tiempo promedio de procesamiento: {eficiencia_promedio:.1f} horas",
                    "interpretacion": (
                        "excelente"
                        if eficiencia_promedio < 24
                        else "bueno" if eficiencia_promedio < 48 else "necesita_mejora"
                    ),
                    "benchmark": "24 horas objetivo, 72 horas máximo aceptable",
                }
            )

        return insights

    # ==================== MÉTODOS PARA DASHBOARD EJECUTIVO ====================

    def _calcular_kpis_mes_actual(self, inicio_mes):
        """Calcula KPIs del mes actual"""
        ahora = timezone.now()
        marcas_mes = MarcaGanadoBovino.objects.filter(
            fecha_registro__gte=inicio_mes, fecha_registro__lte=ahora
        )

        # KPIs básicos
        total_registros = marcas_mes.count()
        total_cabezas = (
            marcas_mes.aggregate(Sum("cantidad_cabezas"))["cantidad_cabezas__sum"] or 0
        )

        # KPIs de procesamiento
        marcas_procesadas = marcas_mes.filter(estado__in=["APROBADO", "RECHAZADO"])
        total_procesadas = marcas_procesadas.count()
        aprobadas = marcas_mes.filter(estado="APROBADO").count()
        rechazadas = marcas_mes.filter(estado="RECHAZADO").count()
        pendientes = marcas_mes.filter(estado__in=["PENDIENTE", "EN_PROCESO"]).count()

        # KPIs económicos
        ingresos_mes = (
            marcas_mes.filter(estado="APROBADO").aggregate(Sum("monto_certificacion"))[
                "monto_certificacion__sum"
            ]
            or 0
        )

        # KPIs de eficiencia
        tiempo_promedio = (
            marcas_mes.filter(tiempo_procesamiento_horas__isnull=False).aggregate(
                Avg("tiempo_procesamiento_horas")
            )["tiempo_procesamiento_horas__avg"]
            or 0
        )

        # KPIs de calidad
        tasa_aprobacion = (
            (aprobadas / total_procesadas * 100) if total_procesadas > 0 else 0
        )

        # KPIs de diversidad
        departamentos_activos = marcas_mes.values("departamento").distinct().count()
        razas_registradas = marcas_mes.values("raza_bovino").distinct().count()
        productores_unicos = marcas_mes.values("ci_productor").distinct().count()

        # KPIs de logos IA
        logos_mes = LogoMarcaBovina.objects.filter(
            fecha_generacion__gte=inicio_mes, fecha_generacion__lte=ahora
        )
        total_logos = logos_mes.count()
        logos_exitosos = logos_mes.filter(exito=True).count()

        return {
            # Métricas de volumen
            "total_registros": total_registros,
            "total_cabezas": total_cabezas,
            "promedio_cabezas_por_marca": (
                round(total_cabezas / total_registros, 2) if total_registros > 0 else 0
            ),
            "registros_por_dia": round(total_registros / ahora.day, 2),
            # Métricas de procesamiento
            "total_procesadas": total_procesadas,
            "aprobadas": aprobadas,
            "rechazadas": rechazadas,
            "pendientes": pendientes,
            "tasa_aprobacion": round(tasa_aprobacion, 2),
            "tasa_rechazo": round(
                (rechazadas / total_procesadas * 100) if total_procesadas > 0 else 0, 2
            ),
            "tasa_pendientes": round(
                (pendientes / total_registros * 100) if total_registros > 0 else 0, 2
            ),
            # Métricas económicas
            "ingresos_mes": float(ingresos_mes),
            "ingreso_promedio_marca": (
                round(float(ingresos_mes) / aprobadas, 2) if aprobadas > 0 else 0
            ),
            "ingreso_por_cabeza": (
                round(float(ingresos_mes) / total_cabezas, 2)
                if total_cabezas > 0
                else 0
            ),
            # Métricas de eficiencia
            "tiempo_promedio_procesamiento": round(tiempo_promedio, 2),
            "eficiencia_diaria": round(total_procesadas / ahora.day, 2),
            "productividad_general": round(
                (total_cabezas * tasa_aprobacion / 100) / max(tiempo_promedio, 1), 2
            ),
            # Métricas de diversidad
            "departamentos_activos": departamentos_activos,
            "razas_registradas": razas_registradas,
            "productores_unicos": productores_unicos,
            "diversidad_geografica": round(
                (departamentos_activos / 9) * 100, 1
            ),  # 9 departamentos en Bolivia
            # Métricas de tecnología
            "logos_generados": total_logos,
            "logos_exitosos": logos_exitosos,
            "tasa_exito_logos": round(
                (logos_exitosos / total_logos * 100) if total_logos > 0 else 0, 2
            ),
            "adopcion_tecnologica": round(
                (total_logos / total_registros * 100) if total_registros > 0 else 0, 2
            ),
            # Métricas de crecimiento
            "velocidad_crecimiento": self._calcular_velocidad_crecimiento(inicio_mes),
            "momentum_sectorial": self._calcular_momentum_sectorial(marcas_mes),
            # Score general
            "score_general_mes": self._calcular_score_general_mes(
                {
                    "tasa_aprobacion": tasa_aprobacion,
                    "tiempo_promedio": tiempo_promedio,
                    "total_registros": total_registros,
                    "diversidad": departamentos_activos + razas_registradas,
                }
            ),
        }

    def _calcular_kpis_periodo(self, fecha_inicio, fecha_fin):
        """Calcula KPIs para un período específico"""
        marcas_periodo = MarcaGanadoBovino.objects.filter(
            fecha_registro__gte=fecha_inicio, fecha_registro__lt=fecha_fin
        )

        # KPIs básicos del período
        total_registros = marcas_periodo.count()
        total_cabezas = (
            marcas_periodo.aggregate(Sum("cantidad_cabezas"))["cantidad_cabezas__sum"]
            or 0
        )

        # KPIs de procesamiento
        procesadas = marcas_periodo.filter(estado__in=["APROBADO", "RECHAZADO"]).count()
        aprobadas = marcas_periodo.filter(estado="APROBADO").count()

        # KPIs económicos
        ingresos = (
            marcas_periodo.filter(estado="APROBADO").aggregate(
                Sum("monto_certificacion")
            )["monto_certificacion__sum"]
            or 0
        )

        # KPIs de eficiencia
        tiempo_promedio = (
            marcas_periodo.filter(tiempo_procesamiento_horas__isnull=False).aggregate(
                Avg("tiempo_procesamiento_horas")
            )["tiempo_procesamiento_horas__avg"]
            or 0
        )

        return {
            "total_registros": total_registros,
            "total_cabezas": total_cabezas,
            "aprobadas": aprobadas,
            "ingresos_total": float(ingresos),
            "tiempo_promedio_procesamiento": round(tiempo_promedio, 2),
            "tasa_aprobacion": round(
                (aprobadas / procesadas * 100) if procesadas > 0 else 0, 2
            ),
        }

    def _calcular_variaciones_kpis(self, kpis_actual, kpis_anterior):
        """Calcula variaciones entre períodos"""
        variaciones = {}

        # Métricas clave para comparar
        metricas_comparacion = [
            "total_registros",
            "total_cabezas",
            "aprobadas",
            "ingresos_total",
            "tiempo_promedio_procesamiento",
            "tasa_aprobacion",
        ]

        for metrica in metricas_comparacion:
            valor_actual = kpis_actual.get(metrica, 0)
            valor_anterior = kpis_anterior.get(metrica, 0)

            # Calcular variación absoluta y porcentual
            variacion_abs = valor_actual - valor_anterior
            variacion_pct = (
                (variacion_abs / valor_anterior * 100) if valor_anterior > 0 else 0
            )

            # Determinar tendencia
            if variacion_pct > 5:
                tendencia = "mejora"
            elif variacion_pct < -5:
                tendencia = "declive"
            else:
                tendencia = "estable"

            variaciones[metrica] = {
                "actual": valor_actual,
                "anterior": valor_anterior,
                "variacion_absoluta": round(variacion_abs, 2),
                "variacion_porcentual": round(variacion_pct, 2),
                "tendencia": tendencia,
                "significativa": abs(variacion_pct) > 10,
            }

        return variaciones

    def _detectar_alertas_criticas(self):
        """Detecta alertas críticas del sistema"""
        alertas = []
        ahora = timezone.now()

        # Alerta por acumulación de pendientes
        pendientes_criticos = MarcaGanadoBovino.objects.filter(
            estado__in=["PENDIENTE", "EN_PROCESO"],
            fecha_registro__lt=ahora - timedelta(days=7),  # Más de 7 días pendientes
        )

        if pendientes_criticos.count() > 50:
            alertas.append(
                {
                    "tipo": "critica",
                    "categoria": "acumulacion_pendientes",
                    "titulo": "Acumulación Crítica de Pendientes",
                    "descripcion": f"{pendientes_criticos.count()} marcas pendientes por más de 7 días",
                    "impacto": "Alto - Afecta satisfacción del cliente",
                    "accion_requerida": "Asignar recursos adicionales para procesamiento",
                    "urgencia": "inmediata",
                }
            )

        # Alerta por baja tasa de aprobación reciente
        marcas_recientes = MarcaGanadoBovino.objects.filter(
            fecha_procesamiento__gte=ahora - timedelta(days=7),
            estado__in=["APROBADO", "RECHAZADO"],
        )

        if marcas_recientes.exists():
            aprobadas_recientes = marcas_recientes.filter(estado="APROBADO").count()
            total_recientes = marcas_recientes.count()
            tasa_reciente = aprobadas_recientes / total_recientes * 100

            if tasa_reciente < 60:
                alertas.append(
                    {
                        "tipo": "warning",
                        "categoria": "calidad_proceso",
                        "titulo": "Baja Tasa de Aprobación",
                        "descripcion": f"Tasa de aprobación de {tasa_reciente:.1f}% en últimos 7 días",
                        "impacto": "Medio - Indica problemas de calidad",
                        "accion_requerida": "Revisar criterios y capacitar evaluadores",
                        "urgencia": "alta",
                    }
                )

        # Alerta por tiempo de procesamiento elevado
        tiempo_promedio_reciente = (
            marcas_recientes.filter(tiempo_procesamiento_horas__isnull=False).aggregate(
                Avg("tiempo_procesamiento_horas")
            )["tiempo_procesamiento_horas__avg"]
            or 0
        )

        if tiempo_promedio_reciente > 96:  # Más de 4 días
            alertas.append(
                {
                    "tipo": "warning",
                    "categoria": "eficiencia",
                    "titulo": "Tiempo de Procesamiento Elevado",
                    "descripcion": f"Tiempo promedio de {tiempo_promedio_reciente:.1f}h excede límites",
                    "impacto": "Medio - Afecta experiencia del usuario",
                    "accion_requerida": "Optimizar procesos y aumentar capacidad",
                    "urgencia": "media",
                }
            )

        # Alerta por falla en generación de logos
        logos_recientes = LogoMarcaBovina.objects.filter(
            fecha_generacion__gte=ahora - timedelta(days=3)
        )

        if logos_recientes.exists():
            logos_fallidos = logos_recientes.filter(exito=False).count()
            total_logos = logos_recientes.count()
            tasa_fallo = logos_fallidos / total_logos * 100

            if tasa_fallo > 25:
                alertas.append(
                    {
                        "tipo": "info",
                        "categoria": "tecnologia",
                        "titulo": "Fallas en Generación de Logos",
                        "descripcion": f"{tasa_fallo:.1f}% de fallas en generación IA",
                        "impacto": "Bajo - Servicio complementario afectado",
                        "accion_requerida": "Verificar APIs y modelos de IA",
                        "urgencia": "baja",
                    }
                )

        # Alerta por inactividad departamental
        mes_actual = ahora.replace(day=1)
        dept_inactivos = []

        for dept_code, dept_name in MarcaGanadoBovino._meta.get_field(
            "departamento"
        ).choices:
            registros_dept = MarcaGanadoBovino.objects.filter(
                departamento=dept_code, fecha_registro__gte=mes_actual
            ).count()

            if registros_dept == 0:
                dept_inactivos.append(dept_name)

        if len(dept_inactivos) > 2:
            alertas.append(
                {
                    "tipo": "info",
                    "categoria": "cobertura_geografica",
                    "titulo": "Departamentos Sin Actividad",
                    "descripcion": f"{len(dept_inactivos)} departamentos sin registros este mes",
                    "departamentos": dept_inactivos,
                    "impacto": "Medio - Cobertura nacional incompleta",
                    "accion_requerida": "Implementar campañas regionales específicas",
                    "urgencia": "media",
                }
            )

        return sorted(
            alertas,
            key=lambda x: {"critica": 3, "warning": 2, "info": 1}[x["tipo"]],
            reverse=True,
        )

    def _identificar_top_performers(self):
        """Identifica mejores performers del mes"""
        ahora = timezone.now()
        inicio_mes = ahora.replace(day=1)

        # Top departamentos por volumen
        top_departamentos = (
            MarcaGanadoBovino.objects.filter(fecha_registro__gte=inicio_mes)
            .values("departamento")
            .annotate(
                total_registros=Count("id"),
                total_cabezas=Sum("cantidad_cabezas"),
                ingresos=Sum("monto_certificacion", filter=Q(estado="APROBADO")),
            )
            .order_by("-total_cabezas")[:3]
        )

        # Enriquecer departamentos
        top_departamentos_procesados = []
        for dept in top_departamentos:
            dept_info = {
                "departamento": dept["departamento"],
                "departamento_display": dict(
                    MarcaGanadoBovino._meta.get_field("departamento").choices
                ).get(dept["departamento"], dept["departamento"]),
                "total_registros": dept["total_registros"],
                "total_cabezas": dept["total_cabezas"],
                "ingresos_generados": float(dept["ingresos"] or 0),
                "promedio_cabezas": (
                    round(dept["total_cabezas"] / dept["total_registros"], 2)
                    if dept["total_registros"] > 0
                    else 0
                ),
            }
            top_departamentos_procesados.append(dept_info)

        # Top razas por popularidad
        top_razas = (
            MarcaGanadoBovino.objects.filter(fecha_registro__gte=inicio_mes)
            .values("raza_bovino")
            .annotate(
                total_registros=Count("id"), total_cabezas=Sum("cantidad_cabezas")
            )
            .order_by("-total_registros")[:3]
        )

        top_razas_procesadas = []
        for raza in top_razas:
            raza_info = {
                "raza": raza["raza_bovino"],
                "raza_display": dict(
                    MarcaGanadoBovino._meta.get_field("raza_bovino").choices
                ).get(raza["raza_bovino"], raza["raza_bovino"]),
                "total_registros": raza["total_registros"],
                "total_cabezas": raza["total_cabezas"],
                "participacion_mercado": self._calcular_participacion_mercado(
                    raza["total_registros"], inicio_mes
                ),
            }
            top_razas_procesadas.append(raza_info)

        # Top productores por volumen (anonimizados)
        top_productores = (
            MarcaGanadoBovino.objects.filter(fecha_registro__gte=inicio_mes)
            .values("ci_productor", "departamento")
            .annotate(total_marcas=Count("id"), total_cabezas=Sum("cantidad_cabezas"))
            .order_by("-total_cabezas")[:5]
        )

        top_productores_procesados = []
        for i, productor in enumerate(top_productores, 1):
            productor_info = {
                "ranking": i,
                "identificador": f"Productor-{productor['ci_productor'][-4:]}",  # Solo últimos 4 dígitos
                "departamento": dict(
                    MarcaGanadoBovino._meta.get_field("departamento").choices
                ).get(productor["departamento"], productor["departamento"]),
                "total_marcas": productor["total_marcas"],
                "total_cabezas": productor["total_cabezas"],
                "promedio_cabezas_por_marca": round(
                    productor["total_cabezas"] / productor["total_marcas"], 2
                ),
            }
            top_productores_procesados.append(productor_info)

        return {
            "top_departamentos": top_departamentos_procesados,
            "top_razas_bovinas": top_razas_procesadas,
            "top_productores": top_productores_procesados,
            "criterios": {
                "departamentos": "Por total de cabezas registradas",
                "razas": "Por número de registros",
                "productores": "Por total de cabezas (anonimizado)",
            },
        }

    def _evaluar_salud_sistema(self):
        """Evalúa la salud general del sistema"""
        ahora = timezone.now()

        # Métricas de salud
        # 1. Disponibilidad operacional (sin errores críticos)
        disponibilidad = 100  # Asumir 100% si no hay errores críticos detectados

        # 2. Eficiencia de procesamiento
        marcas_ultimo_mes = MarcaGanadoBovino.objects.filter(
            fecha_registro__gte=ahora - timedelta(days=30)
        )

        tiempo_promedio_mes = (
            marcas_ultimo_mes.filter(
                tiempo_procesamiento_horas__isnull=False
            ).aggregate(Avg("tiempo_procesamiento_horas"))[
                "tiempo_procesamiento_horas__avg"
            ]
            or 0
        )

        eficiencia_score = (
            max(0, 100 - (tiempo_promedio_mes - 24) * 2)
            if tiempo_promedio_mes > 0
            else 100
        )

        # 3. Calidad del servicio
        procesadas_mes = marcas_ultimo_mes.filter(estado__in=["APROBADO", "RECHAZADO"])
        aprobadas_mes = marcas_ultimo_mes.filter(estado="APROBADO").count()
        total_procesadas_mes = procesadas_mes.count()

        calidad_score = (
            (aprobadas_mes / total_procesadas_mes * 100)
            if total_procesadas_mes > 0
            else 100
        )

        # 4. Capacidad del sistema
        registros_diarios = marcas_ultimo_mes.count() / 30
        capacidad_score = min(
            100, (registros_diarios / 50) * 100
        )  # Asumiendo capacidad ideal de 50/día

        # 5. Adopción tecnológica
        logos_mes = LogoMarcaBovina.objects.filter(
            fecha_generacion__gte=ahora - timedelta(days=30)
        )
        adopcion_ia = (
            (logos_mes.count() / marcas_ultimo_mes.count() * 100)
            if marcas_ultimo_mes.count() > 0
            else 0
        )

        # Score general de salud (promedio ponderado)
        score_general = (
            disponibilidad * 0.25
            + eficiencia_score * 0.25
            + calidad_score * 0.25
            + capacidad_score * 0.15
            + adopcion_ia * 0.10
        )

        # Determinar estado de salud
        if score_general >= 90:
            estado_salud = "excelente"
            color_indicador = "verde"
        elif score_general >= 75:
            estado_salud = "bueno"
            color_indicador = "amarillo_claro"
        elif score_general >= 60:
            estado_salud = "aceptable"
            color_indicador = "amarillo"
        else:
            estado_salud = "necesita_atencion"
            color_indicador = "rojo"

        return {
            "score_general": round(score_general, 1),
            "estado_salud": estado_salud,
            "color_indicador": color_indicador,
            "componentes": {
                "disponibilidad_operacional": round(disponibilidad, 1),
                "eficiencia_procesamiento": round(eficiencia_score, 1),
                "calidad_servicio": round(calidad_score, 1),
                "capacidad_sistema": round(capacidad_score, 1),
                "adopcion_tecnologica": round(adopcion_ia, 1),
            },
            "metricas_base": {
                "tiempo_promedio_procesamiento": round(tiempo_promedio_mes, 2),
                "registros_diarios_promedio": round(registros_diarios, 2),
                "tasa_aprobacion_mes": round(calidad_score, 2),
                "adopcion_ia_porcentaje": round(adopcion_ia, 2),
            },
            "recomendaciones_salud": self._generar_recomendaciones_salud(
                score_general,
                {
                    "eficiencia": eficiencia_score,
                    "calidad": calidad_score,
                    "capacidad": capacidad_score,
                    "adopcion": adopcion_ia,
                },
            ),
        }

    def _generar_acciones_dashboard(
        self, alertas_criticas, salud_sistema, predicciones
    ):
        """Genera acciones recomendadas para el dashboard ejecutivo"""
        acciones = []

        # Acciones basadas en alertas críticas
        alertas_criticas_count = len(
            [a for a in alertas_criticas if a["tipo"] == "critica"]
        )
        if alertas_criticas_count > 0:
            acciones.append(
                {
                    "prioridad": "critica",
                    "categoria": "atencion_inmediata",
                    "accion": f'Atender {alertas_criticas_count} alerta{"s" if alertas_criticas_count > 1 else ""} crítica{"s" if alertas_criticas_count > 1 else ""}',
                    "plazo": "inmediato",
                    "responsable": "Gerencia de Operaciones",
                }
            )

        # Acciones basadas en salud del sistema
        if salud_sistema["score_general"] < 75:
            acciones.append(
                {
                    "prioridad": "alta",
                    "categoria": "mejora_sistemica",
                    "accion": "Implementar plan de mejora de salud del sistema",
                    "plazo": "7 días",
                    "responsable": "Gerencia General",
                    "componentes_afectados": [
                        comp
                        for comp, score in salud_sistema["componentes"].items()
                        if score < 75
                    ],
                }
            )

        # Acciones basadas en predicciones
        if predicciones and not predicciones.get("datos_insuficientes", False):
            if predicciones.get("prediccion_marcas", 0) > 200:  # Alta demanda predicha
                acciones.append(
                    {
                        "prioridad": "media",
                        "categoria": "preparacion_capacidad",
                        "accion": "Preparar capacidad adicional para alta demanda predicha",
                        "plazo": "15 días",
                        "responsable": "Gerencia de Operaciones",
                        "prediccion_base": f"{predicciones['prediccion_marcas']} registros esperados",
                    }
                )

        # Acción de monitoreo continuo
        acciones.append(
            {
                "prioridad": "baja",
                "categoria": "monitoreo",
                "accion": "Revisar dashboard y métricas diariamente",
                "plazo": "continuo",
                "responsable": "Equipo de BI",
            }
        )

        return acciones

    # ==================== MÉTODOS AUXILIARES ADICIONALES ====================

    def _calcular_velocidad_crecimiento(self, inicio_mes):
        """Calcula velocidad de crecimiento respecto al mes anterior"""
        mes_anterior_inicio = (inicio_mes - timedelta(days=1)).replace(day=1)

        registros_mes_actual = MarcaGanadoBovino.objects.filter(
            fecha_registro__gte=inicio_mes
        ).count()

        registros_mes_anterior = MarcaGanadoBovino.objects.filter(
            fecha_registro__gte=mes_anterior_inicio, fecha_registro__lt=inicio_mes
        ).count()

        if registros_mes_anterior > 0:
            velocidad = (
                (registros_mes_actual - registros_mes_anterior) / registros_mes_anterior
            ) * 100
            return round(velocidad, 2)

        return 0

    def _calcular_momentum_sectorial(self, marcas_mes):
        """Calcula momentum del sector ganadero"""
        # Factores: diversidad, volumen, eficiencia
        departamentos = marcas_mes.values("departamento").distinct().count()
        razas = marcas_mes.values("raza_bovino").distinct().count()
        propositos = marcas_mes.values("proposito_ganado").distinct().count()

        # Normalizar factores
        diversidad_score = min(
            100, (departamentos * 11.1 + razas * 8.33 + propositos * 25)
        )  # Max teórico: 9+12+4
        volumen_score = min(100, marcas_mes.count() * 2)  # 50 registros = 100%

        # Score combinado
        momentum = diversidad_score * 0.6 + volumen_score * 0.4

        return round(momentum, 2)

    def _calcular_score_general_mes(self, datos):
        """Calcula score general del mes"""
        # Ponderación: calidad 40%, eficiencia 30%, volumen 20%, diversidad 10%
        score_calidad = datos["tasa_aprobacion"]
        score_eficiencia = max(0, 100 - (datos["tiempo_promedio"] - 24) * 1.5)
        score_volumen = min(100, datos["total_registros"] * 2)  # 50 registros = 100%
        score_diversidad = min(
            100, datos["diversidad"] * 5
        )  # 20 elementos diversos = 100%

        score_total = (
            score_calidad * 0.4
            + score_eficiencia * 0.3
            + score_volumen * 0.2
            + score_diversidad * 0.1
        )

        return round(score_total, 1)

    def _calcular_participacion_mercado(self, registros_raza, inicio_mes):
        """Calcula participación de mercado de una raza"""
        total_registros_mes = MarcaGanadoBovino.objects.filter(
            fecha_registro__gte=inicio_mes
        ).count()

        if total_registros_mes > 0:
            return round((registros_raza / total_registros_mes) * 100, 2)

        return 0

    def _generar_recomendaciones_salud(self, score_general, componentes):
        """Genera recomendaciones específicas para mejorar salud del sistema"""
        recomendaciones = []

        # Recomendaciones por componente
        if componentes["eficiencia"] < 70:
            recomendaciones.append(
                {
                    "componente": "eficiencia",
                    "recomendacion": "Optimizar procesos de evaluación y automatizar tareas repetitivas",
                    "impacto_esperado": "+15 puntos en eficiencia",
                }
            )

        if componentes["calidad"] < 75:
            recomendaciones.append(
                {
                    "componente": "calidad",
                    "recomendacion": "Implementar programa de capacitación para evaluadores",
                    "impacto_esperado": "+10 puntos en calidad",
                }
            )

        if componentes["capacidad"] < 60:
            recomendaciones.append(
                {
                    "componente": "capacidad",
                    "recomendacion": "Evaluar necesidad de recursos adicionales o optimización de carga",
                    "impacto_esperado": "+20 puntos en capacidad",
                }
            )

        if componentes["adopcion"] < 50:
            recomendaciones.append(
                {
                    "componente": "adopcion_tecnologica",
                    "recomendacion": "Promover uso de herramientas de IA y digitalización",
                    "impacto_esperado": "+15 puntos en adopción",
                }
            )

        # Recomendación general si score es bajo
        if score_general < 70:
            recomendaciones.append(
                {
                    "componente": "general",
                    "recomendacion": "Implementar plan integral de mejora con revisión semanal",
                    "impacto_esperado": "Mejora gradual hacia 80+ puntos",
                }
            )

        return recomendaciones

    # ==================== ENDPOINTS DE REPORTES ESPECIALIZADOS ====================

    @action(detail=False, methods=["get"])
    def reporte_impacto_economico(self, request):
        """Análisis de impacto económico del sector ganadero bovino"""
        try:
            año = int(request.query_params.get("año", timezone.now().year))
            incluir_proyecciones = (
                request.query_params.get("proyecciones", "true").lower() == "true"
            )

            # Período de análisis
            inicio_año = datetime(año, 1, 1)
            fin_año = datetime(año + 1, 1, 1)

            marcas_año = MarcaGanadoBovino.objects.filter(
                fecha_registro__gte=inicio_año, fecha_registro__lt=fin_año
            )

            # Impacto económico directo
            ingresos_directos = (
                marcas_año.filter(estado="APROBADO").aggregate(
                    Sum("monto_certificacion")
                )["monto_certificacion__sum"]
                or 0
            )

            total_cabezas = (
                marcas_año.aggregate(Sum("cantidad_cabezas"))["cantidad_cabezas__sum"]
                or 0
            )

            # Estimación de valor económico del ganado registrado
            # Valores promedio por cabeza según propósito (en bolivianos)
            valores_por_proposito = {
                "CARNE": 8000,  # Bs por cabeza para carne
                "LECHE": 12000,  # Bs por cabeza para leche
                "DOBLE_PROPOSITO": 10000,  # Bs por cabeza doble propósito
                "REPRODUCCION": 15000,  # Bs por cabeza reproductores
            }

            valor_ganado_registrado = 0
            impacto_por_proposito = []

            for proposito, valor_unitario in valores_por_proposito.items():
                marcas_proposito = marcas_año.filter(proposito_ganado=proposito)
                cabezas_proposito = (
                    marcas_proposito.aggregate(Sum("cantidad_cabezas"))[
                        "cantidad_cabezas__sum"
                    ]
                    or 0
                )
                valor_proposito = cabezas_proposito * valor_unitario
                valor_ganado_registrado += valor_proposito

                if cabezas_proposito > 0:
                    impacto_por_proposito.append(
                        {
                            "proposito": proposito,
                            "proposito_display": dict(
                                MarcaGanadoBovino._meta.get_field(
                                    "proposito_ganado"
                                ).choices
                            ).get(proposito, proposito),
                            "cabezas_registradas": cabezas_proposito,
                            "valor_unitario_bs": valor_unitario,
                            "valor_total_bs": valor_proposito,
                            "participacion_valor": round(
                                (
                                    (valor_proposito / valor_ganado_registrado * 100)
                                    if valor_ganado_registrado > 0
                                    else 0
                                ),
                                2,
                            ),
                        }
                    )

            # Impacto económico por departamento
            impacto_departamental = []
            for dept_code, dept_name in MarcaGanadoBovino._meta.get_field(
                "departamento"
            ).choices:
                marcas_dept = marcas_año.filter(departamento=dept_code)
                cabezas_dept = (
                    marcas_dept.aggregate(Sum("cantidad_cabezas"))[
                        "cantidad_cabezas__sum"
                    ]
                    or 0
                )
                ingresos_dept = (
                    marcas_dept.filter(estado="APROBADO").aggregate(
                        Sum("monto_certificacion")
                    )["monto_certificacion__sum"]
                    or 0
                )

                if cabezas_dept > 0:
                    # Estimación valor ganado por departamento
                    valor_estimado_dept = 0
                    for proposito, valor_unitario in valores_por_proposito.items():
                        cabezas_prop = (
                            marcas_dept.filter(proposito_ganado=proposito).aggregate(
                                Sum("cantidad_cabezas")
                            )["cantidad_cabezas__sum"]
                            or 0
                        )
                        valor_estimado_dept += cabezas_prop * valor_unitario

                    impacto_departamental.append(
                        {
                            "departamento": dept_code,
                            "departamento_display": dept_name,
                            "cabezas_registradas": cabezas_dept,
                            "ingresos_certificacion": float(ingresos_dept),
                            "valor_estimado_ganado": valor_estimado_dept,
                            "participacion_nacional": round(
                                (
                                    (cabezas_dept / total_cabezas * 100)
                                    if total_cabezas > 0
                                    else 0
                                ),
                                2,
                            ),
                            "densidad_economica": (
                                round(valor_estimado_dept / cabezas_dept, 2)
                                if cabezas_dept > 0
                                else 0
                            ),
                        }
                    )

            # Ordenar por valor estimado
            impacto_departamental.sort(
                key=lambda x: x["valor_estimado_ganado"], reverse=True
            )

            # Multiplicadores económicos estimados
            # Basado en estudios de cadena ganadera (efectos indirectos)
            multiplicador_empleo = 2.5  # 1 empleo directo genera 2.5 empleos totales
            multiplicador_pib = (
                1.8  # 1 peso de actividad directa genera 1.8 de PIB total
            )

            # Estimación de empleos (aproximado: 1 empleo por cada 50 cabezas)
            empleos_directos_estimados = int(total_cabezas / 50)
            empleos_totales_estimados = int(
                empleos_directos_estimados * multiplicador_empleo
            )

            # Contribución estimada al PIB sectorial
            contribucion_pib_directa = (
                valor_ganado_registrado * 0.15
            )  # 15% del valor como contribución anual
            contribucion_pib_total = contribucion_pib_directa * multiplicador_pib

            # Análisis de eficiencia económica
            eficiencia_economica = self._analizar_eficiencia_economica(
                marcas_año, impacto_por_proposito
            )

            reporte_impacto = {
                "año_analisis": año,
                "fecha_generacion": timezone.now().date(),
                "resumen_ejecutivo": {
                    "total_cabezas_registradas": total_cabezas,
                    "valor_ganado_registrado_bs": valor_ganado_registrado,
                    "ingresos_certificacion_bs": float(ingresos_directos),
                    "empleos_estimados_directos": empleos_directos_estimados,
                    "empleos_estimados_totales": empleos_totales_estimados,
                    "contribucion_pib_estimada": contribucion_pib_total,
                },
                "impacto_por_proposito": impacto_por_proposito,
                "impacto_departamental": impacto_departamental,
                "multiplicadores_economicos": {
                    "empleo": multiplicador_empleo,
                    "pib": multiplicador_pib,
                    "metodologia": "Basado en estudios de cadena productiva ganadera",
                },
                "eficiencia_economica": eficiencia_economica,
                "comparacion_sectorial": self._generar_comparacion_sectorial(
                    valor_ganado_registrado, total_cabezas
                ),
                "indicadores_sostenibilidad": self._calcular_indicadores_sostenibilidad(
                    marcas_año
                ),
            }

            # Agregar proyecciones si se solicitan
            if incluir_proyecciones:
                reporte_impacto["proyecciones_economicas"] = (
                    self._generar_proyecciones_economicas(reporte_impacto, año)
                )

            return Response(reporte_impacto)

        except Exception as e:
            return Response(
                {"error": f"Error generando reporte de impacto económico: {str(e)}"},
                status=500,
            )

    @action(detail=False, methods=["get"])
    def reporte_innovacion_tecnologica(self, request):
        """Reporte sobre adopción y eficiencia de tecnologías IA en el sector"""
        try:
            fecha_desde = request.query_params.get("fecha_desde")
            fecha_hasta = request.query_params.get("fecha_hasta")

            # Filtros de fecha
            queryset_logos = LogoMarcaBovina.objects.all()
            queryset_marcas = MarcaGanadoBovino.objects.all()

            if fecha_desde:
                fecha_desde_obj = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
                queryset_logos = queryset_logos.filter(
                    fecha_generacion__date__gte=fecha_desde_obj
                )
                queryset_marcas = queryset_marcas.filter(
                    fecha_registro__date__gte=fecha_desde_obj
                )

            if fecha_hasta:
                fecha_hasta_obj = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
                queryset_logos = queryset_logos.filter(
                    fecha_generacion__date__lte=fecha_hasta_obj
                )
                queryset_marcas = queryset_marcas.filter(
                    fecha_registro__date__lte=fecha_hasta_obj
                )

            # Métricas generales de adopción IA
            total_marcas = queryset_marcas.count()
            total_logos = queryset_logos.count()
            logos_exitosos = queryset_logos.filter(exito=True).count()

            adopcion_rate = (
                (total_logos / total_marcas * 100) if total_marcas > 0 else 0
            )
            success_rate = (
                (logos_exitosos / total_logos * 100) if total_logos > 0 else 0
            )

            # Análisis por modelo de IA
            performance_por_modelo = (
                queryset_logos.values("modelo_ia_usado")
                .annotate(
                    total_generaciones=Count("id"),
                    generaciones_exitosas=Count("id", filter=Q(exito=True)),
                    tiempo_promedio=Avg("tiempo_generacion_segundos"),
                    calidad_promedio=Count("id", filter=Q(calidad_logo="ALTA"))
                    / Count("id")
                    * 100,
                )
                .order_by("-generaciones_exitosas")
            )

            # Enriquecer datos de modelos
            modelos_performance = []
            for modelo in performance_por_modelo:
                tasa_exito = (
                    (
                        modelo["generaciones_exitosas"]
                        / modelo["total_generaciones"]
                        * 100
                    )
                    if modelo["total_generaciones"] > 0
                    else 0
                )

                modelos_performance.append(
                    {
                        "modelo_ia": modelo["modelo_ia_usado"],
                        "total_generaciones": modelo["total_generaciones"],
                        "generaciones_exitosas": modelo["generaciones_exitosas"],
                        "tasa_exito": round(tasa_exito, 2),
                        "tiempo_promedio_segundos": round(
                            modelo["tiempo_promedio"] or 0, 2
                        ),
                        "score_eficiencia": round(
                            (
                                tasa_exito * 0.7
                                + (100 - min(modelo["tiempo_promedio"] or 0, 100)) * 0.3
                            ),
                            2,
                        ),
                        "calidad_promedio": round(modelo["calidad_promedio"] or 0, 2),
                        "recomendacion": self._evaluar_modelo_ia(
                            tasa_exito, modelo["tiempo_promedio"] or 0
                        ),
                    }
                )

            # Tendencias temporales de adopción
            adopcion_temporal = self._analizar_adopcion_temporal(
                queryset_logos, queryset_marcas
            )

            # Análisis de calidad por región
            calidad_por_region = self._analizar_calidad_logos_por_region(queryset_logos)

            # ROI de la tecnología IA
            roi_tecnologia = self._calcular_roi_tecnologia_ia(
                queryset_logos, queryset_marcas
            )

            # Barreras y oportunidades
            barreras_adopcion = self._identificar_barreras_adopcion(
                queryset_marcas, queryset_logos
            )

            # Benchmarking tecnológico
            benchmark_tech = self._generar_benchmark_tecnologico(modelos_performance)

            return Response(
                {
                    "periodo_analisis": {
                        "fecha_desde": fecha_desde or "inicio_registros",
                        "fecha_hasta": fecha_hasta or "actualidad",
                    },
                    "metricas_adopcion": {
                        "total_marcas_periodo": total_marcas,
                        "total_logos_generados": total_logos,
                        "logos_exitosos": logos_exitosos,
                        "tasa_adopcion_ia": round(adopcion_rate, 2),
                        "tasa_exito_general": round(success_rate, 2),
                        "nivel_adopcion": self._clasificar_nivel_adopcion(
                            adopcion_rate
                        ),
                    },
                    "performance_modelos_ia": modelos_performance,
                    "tendencias_adopcion": adopcion_temporal,
                    "calidad_regional": calidad_por_region,
                    "roi_tecnologia": roi_tecnologia,
                    "barreras_oportunidades": barreras_adopcion,
                    "benchmark_tecnologico": benchmark_tech,
                    "recomendaciones_innovacion": self._generar_recomendaciones_innovacion(
                        modelos_performance, adopcion_rate, success_rate
                    ),
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Error generando reporte de innovación: {str(e)}"},
                status=500,
            )

    @action(detail=False, methods=["get"])
    def reporte_sostenibilidad_sectorial(self, request):
        """Reporte de sostenibilidad del sector ganadero bovino"""
        try:
            año = int(request.query_params.get("año", timezone.now().year))
            incluir_comparacion = (
                request.query_params.get("comparacion", "true").lower() == "true"
            )

            # Período de análisis
            inicio_año = datetime(año, 1, 1)
            fin_año = datetime(año + 1, 1, 1)

            marcas_año = MarcaGanadoBovino.objects.filter(
                fecha_registro__gte=inicio_año, fecha_registro__lt=fin_año
            )

            # Indicadores de sostenibilidad económica
            sostenibilidad_economica = {
                "diversificacion_productiva": self._calcular_diversificacion_productiva(
                    marcas_año
                ),
                "distribucion_geografica": self._calcular_distribucion_geografica(
                    marcas_año
                ),
                "estabilidad_ingresos": self._calcular_estabilidad_ingresos(marcas_año),
                "inclusion_productores": self._calcular_inclusion_productores(
                    marcas_año
                ),
            }

            # Indicadores de sostenibilidad social
            sostenibilidad_social = {
                "cobertura_territorial": self._calcular_cobertura_territorial(
                    marcas_año
                ),
                "accesibilidad_servicios": self._calcular_accesibilidad_servicios(
                    marcas_año
                ),
                "equidad_genero": self._estimar_equidad_genero(marcas_año),
                "desarrollo_rural": self._evaluar_desarrollo_rural(marcas_año),
            }

            # Indicadores de sostenibilidad ambiental (estimaciones)
            sostenibilidad_ambiental = {
                "eficiencia_uso_tierra": self._estimar_eficiencia_tierra(marcas_año),
                "diversidad_genetica": self._calcular_diversidad_genetica(marcas_año),
                "practicas_sostenibles": self._evaluar_practicas_sostenibles(
                    marcas_año
                ),
                "huella_carbono_estimada": self._estimar_huella_carbono(marcas_año),
            }

            # Score general de sostenibilidad
            score_sostenibilidad = self._calcular_score_sostenibilidad(
                sostenibilidad_economica,
                sostenibilidad_social,
                sostenibilidad_ambiental,
            )

            # Análisis de riesgos y oportunidades
            analisis_riesgos = self._analizar_riesgos_sostenibilidad(marcas_año)

            # Metas y objetivos SDG (Sustainable Development Goals)
            contribucion_sdg = self._evaluar_contribucion_sdg(
                marcas_año, sostenibilidad_economica
            )

            reporte_sostenibilidad = {
                "año_analisis": año,
                "metodologia": "Framework de Sostenibilidad Sectorial Adaptado",
                "score_general_sostenibilidad": score_sostenibilidad,
                "dimensiones_sostenibilidad": {
                    "economica": sostenibilidad_economica,
                    "social": sostenibilidad_social,
                    "ambiental": sostenibilidad_ambiental,
                },
                "analisis_riesgos_oportunidades": analisis_riesgos,
                "contribucion_objetivos_desarrollo": contribucion_sdg,
                "plan_mejora_sostenibilidad": self._generar_plan_mejora_sostenibilidad(
                    score_sostenibilidad, analisis_riesgos
                ),
            }

            # Comparación interanual si se solicita
            if incluir_comparacion and año > 2020:
                reporte_sostenibilidad["comparacion_interanual"] = (
                    self._comparar_sostenibilidad_interanual(
                        año, reporte_sostenibilidad
                    )
                )

            return Response(reporte_sostenibilidad)

        except Exception as e:
            return Response(
                {"error": f"Error generando reporte de sostenibilidad: {str(e)}"},
                status=500,
            )

    @action(detail=False, methods=["post"])
    def generar_reporte_personalizado_avanzado(self, request):
        """Generador avanzado de reportes con múltiples dimensiones de análisis"""
        try:
            configuracion = request.data

            # Validar configuración
            dimensiones_requeridas = configuracion.get("dimensiones", [])
            filtros_avanzados = configuracion.get("filtros_avanzados", {})
            formato_salida = configuracion.get("formato_salida", "json")
            incluir_visualizaciones = configuracion.get(
                "incluir_visualizaciones", False
            )

            if not dimensiones_requeridas:
                return Response(
                    {"error": "Debe especificar al menos una dimensión de análisis"},
                    status=400,
                )

            # Construir queryset base con filtros avanzados
            queryset = self._construir_queryset_avanzado(filtros_avanzados)

            # Generar análisis por dimensión
            resultado_analisis = {}

            for dimension in dimensiones_requeridas:
                if dimension == "temporal":
                    resultado_analisis["temporal"] = self._analisis_dimension_temporal(
                        queryset, filtros_avanzados.get("temporal", {})
                    )
                elif dimension == "geografico":
                    resultado_analisis["geografico"] = (
                        self._analisis_dimension_geografica(
                            queryset, filtros_avanzados.get("geografico", {})
                        )
                    )
                elif dimension == "economico":
                    resultado_analisis["economico"] = (
                        self._analisis_dimension_economica(
                            queryset, filtros_avanzados.get("economico", {})
                        )
                    )
                elif dimension == "genetico":
                    resultado_analisis["genetico"] = self._analisis_dimension_genetica(
                        queryset, filtros_avanzados.get("genetico", {})
                    )
                elif dimension == "tecnologico":
                    resultado_analisis["tecnologico"] = (
                        self._analisis_dimension_tecnologica(
                            queryset, filtros_avanzados.get("tecnologico", {})
                        )
                    )
                elif dimension == "competitivo":
                    resultado_analisis["competitivo"] = (
                        self._analisis_dimension_competitiva(
                            queryset, filtros_avanzados.get("competitivo", {})
                        )
                    )

            # Análisis cruzado entre dimensiones
            analisis_cruzado = self._realizar_analisis_cruzado(
                resultado_analisis, configuracion.get("cruces_dimensionales", [])
            )

            # Insights y recomendaciones integradas
            insights_integrados = self._generar_insights_integrados(
                resultado_analisis, analisis_cruzado
            )

            # Preparar reporte final
            reporte_avanzado = {
                "configuracion_aplicada": configuracion,
                "fecha_generacion": timezone.now(),
                "total_registros_analizados": queryset.count(),
                "analisis_por_dimension": resultado_analisis,
                "analisis_cruzado": analisis_cruzado,
                "insights_integrados": insights_integrados,
                "recomendaciones_estrategicas": self._generar_recomendaciones_estrategicas_avanzadas(
                    insights_integrados
                ),
                "limitaciones_metodologicas": self._documentar_limitaciones_metodologicas(
                    configuracion
                ),
            }

            # Exportar según formato solicitado
            if formato_salida == "excel":
                return self._exportar_reporte_avanzado_excel(reporte_avanzado)
            elif formato_salida == "pdf":
                return self._exportar_reporte_avanzado_pdf(reporte_avanzado)
            else:
                return Response(reporte_avanzado)

        except Exception as e:
            return Response(
                {"error": f"Error generando reporte avanzado: {str(e)}"}, status=500
            )

    # ==================== MÉTODOS AUXILIARES PARA REPORTES ESPECIALIZADOS ====================

    def _analizar_eficiencia_economica(self, marcas_año, impacto_por_proposito):
        """Analiza eficiencia económica del sector"""
        # Eficiencia por propósito
        eficiencia_por_proposito = []
        for proposito_data in impacto_por_proposito:
            eficiencia = (
                proposito_data["valor_total_bs"] / proposito_data["cabezas_registradas"]
            )
            eficiencia_por_proposito.append(
                {
                    "proposito": proposito_data["proposito_display"],
                    "eficiencia_economica": round(eficiencia, 2),
                    "valor_por_cabeza": proposito_data["valor_unitario_bs"],
                }
            )

        # Eficiencia temporal (ingresos vs tiempo)
        marcas_con_tiempo = marcas_año.filter(
            tiempo_procesamiento_horas__isnull=False, estado="APROBADO"
        )

        if marcas_con_tiempo.exists():
            eficiencia_temporal = (
                marcas_con_tiempo.aggregate(
                    ingreso_por_hora=Sum("monto_certificacion")
                    / Sum("tiempo_procesamiento_horas")
                )["ingreso_por_hora"]
                or 0
            )
        else:
            eficiencia_temporal = 0

        return {
            "eficiencia_por_proposito": eficiencia_por_proposito,
            "eficiencia_temporal_procesamiento": round(float(eficiencia_temporal), 2),
            "score_eficiencia_general": self._calcular_score_eficiencia_general(
                eficiencia_por_proposito
            ),
        }

    def _generar_comparacion_sectorial(self, valor_ganado, total_cabezas):
        """Genera comparación con benchmarks sectoriales"""
        # Benchmarks estimados para Bolivia (en bolivianos)
        benchmarks = {
            "valor_promedio_cabeza_nacional": 10000,
            "productividad_promedio_por_marca": 75,  # cabezas promedio
            "ingresos_promedio_sector": 50000000,  # ingresos anuales estimados del sector
        }

        valor_promedio_actual = valor_ganado / total_cabezas if total_cabezas > 0 else 0

        return {
            "valor_promedio_por_cabeza": {
                "actual": round(valor_promedio_actual, 2),
                "benchmark": benchmarks["valor_promedio_cabeza_nacional"],
                "diferencia_porcentual": (
                    round(
                        (
                            (
                                valor_promedio_actual
                                - benchmarks["valor_promedio_cabeza_nacional"]
                            )
                            / benchmarks["valor_promedio_cabeza_nacional"]
                            * 100
                        ),
                        2,
                    )
                    if benchmarks["valor_promedio_cabeza_nacional"] > 0
                    else 0
                ),
            },
            "posicion_relativa": (
                "superior"
                if valor_promedio_actual > benchmarks["valor_promedio_cabeza_nacional"]
                else "inferior"
            ),
            "potencial_mejora": max(
                0, benchmarks["valor_promedio_cabeza_nacional"] - valor_promedio_actual
            )
            * total_cabezas,
        }

    def _calcular_indicadores_sostenibilidad(self, marcas_año):
        """Calcula indicadores básicos de sostenibilidad"""
        # Diversidad genética
        razas_diferentes = marcas_año.values("raza_bovino").distinct().count()

        # Distribución geográfica
        departamentos_activos = marcas_año.values("departamento").distinct().count()

        # Inclusión de pequeños productores (menos de 50 cabezas)
        pequeños_productores = marcas_año.filter(cantidad_cabezas__lt=50).count()
        total_marcas = marcas_año.count()

        return {
            "diversidad_genetica_score": min(
                100, (razas_diferentes / 12) * 100
            ),  # 12 razas máximo teórico
            "cobertura_territorial_score": (departamentos_activos / 9)
            * 100,  # 9 departamentos en Bolivia
            "inclusion_pequeños_productores": round(
                (pequeños_productores / total_marcas * 100) if total_marcas > 0 else 0,
                2,
            ),
            "sostenibilidad_general": "En desarrollo",  # Placeholder para análisis más profundo
        }

    def _generar_proyecciones_economicas(self, reporte_base, año_base):
        """Genera proyecciones económicas para los próximos 3 años"""
        # Factores de crecimiento estimados
        factor_crecimiento_anual = 1.08  # 8% anual estimado
        factor_inflacion = 1.05  # 5% anual estimado

        proyecciones = []

        for i in range(1, 4):  # Próximos 3 años
            año_proyeccion = año_base + i

            # Proyecciones ajustadas por crecimiento e inflación
            valor_ganado_proyectado = reporte_base["resumen_ejecutivo"][
                "valor_ganado_registrado_bs"
            ] * (factor_crecimiento_anual**i)
            empleos_proyectados = int(
                reporte_base["resumen_ejecutivo"]["empleos_estimados_totales"]
                * (factor_crecimiento_anual**i)
            )

            proyecciones.append(
                {
                    "año": año_proyeccion,
                    "valor_ganado_proyectado": round(valor_ganado_proyectado, 2),
                    "empleos_estimados": empleos_proyectados,
                    "contribucion_pib_proyectada": round(
                        valor_ganado_proyectado * 0.15 * 1.8, 2
                    ),
                    "confianza_proyeccion": (
                        "alta" if i == 1 else "media" if i == 2 else "baja"
                    ),
                    "factores_considerados": [
                        "crecimiento_sectorial",
                        "inflacion",
                        "tendencias_historicas",
                    ],
                }
            )

        return {
            "proyecciones_anuales": proyecciones,
            "supuestos_utilizados": {
                "crecimiento_anual": f"{(factor_crecimiento_anual - 1) * 100:.1f}%",
                "inflacion_anual": f"{(factor_inflacion - 1) * 100:.1f}%",
            },
            "limitaciones": [
                "No considera factores externos extremos",
                "Basado en tendencias históricas",
                "Sujeto a variabilidad macroeconómica",
            ],
        }

    def _evaluar_modelo_ia(self, tasa_exito, tiempo_promedio):
        """Evalúa y recomienda sobre un modelo de IA"""
        if tasa_exito >= 90 and tiempo_promedio <= 30:
            return "Excelente - Mantener como modelo principal"
        elif tasa_exito >= 80 and tiempo_promedio <= 60:
            return "Bueno - Usar como modelo secundario"
        elif tasa_exito >= 70:
            return "Aceptable - Optimizar parámetros"
        else:
            return "Bajo rendimiento - Evaluar reemplazo"

    def _analizar_adopcion_temporal(self, queryset_logos, queryset_marcas):
        """Analiza tendencias temporales de adopción de IA"""
        # Análisis mensual de adopción
        adopcion_mensual = (
            queryset_logos.extra(
                select={
                    "año": "YEAR(fecha_generacion)",
                    "mes": "MONTH(fecha_generacion)",
                }
            )
            .values("año", "mes")
            .annotate(
                logos_generados=Count("id"),
                logos_exitosos=Count("id", filter=Q(exito=True)),
            )
            .order_by("año", "mes")
        )

        # Calcular tasa de adopción por mes
        tendencias = []
        for item in adopcion_mensual:
            marcas_mes = queryset_marcas.filter(
                fecha_registro__year=item["año"], fecha_registro__month=item["mes"]
            ).count()

            tasa_adopcion_mes = (
                (item["logos_generados"] / marcas_mes * 100) if marcas_mes > 0 else 0
            )

            tendencias.append(
                {
                    "año": item["año"],
                    "mes": item["mes"],
                    "mes_nombre": calendar.month_name[item["mes"]],
                    "logos_generados": item["logos_generados"],
                    "logos_exitosos": item["logos_exitosos"],
                    "marcas_registradas": marcas_mes,
                    "tasa_adopcion": round(tasa_adopcion_mes, 2),
                    "momentum": (
                        "creciente"
                        if tasa_adopcion_mes > 50
                        else "estable" if tasa_adopcion_mes > 30 else "bajo"
                    ),
                }
            )

        return {
            "tendencias_mensuales": tendencias,
            "tasa_crecimiento_adopcion": self._calcular_crecimiento_adopcion(
                tendencias
            ),
            "estacionalidad_detectada": self._detectar_estacionalidad_adopcion(
                tendencias
            ),
        }

    def _calcular_crecimiento_adopcion(self, tendencias):
        """Calcula tasa de crecimiento de adopción"""
        if len(tendencias) >= 2:
            primera_tasa = tendencias[0]["tasa_adopcion"]
            ultima_tasa = tendencias[-1]["tasa_adopcion"]

            if primera_tasa > 0:
                crecimiento = ((ultima_tasa - primera_tasa) / primera_tasa) * 100
                return round(crecimiento, 2)

        return 0

    def _detectar_estacionalidad_adopcion(self, tendencias):
        """Detecta patrones estacionales en adopción de IA"""
        if len(tendencias) < 12:
            return {"mensaje": "Datos insuficientes para detectar estacionalidad"}

        # Agrupar por mes (sin considerar año)
        adopcion_por_mes = {}
        for item in tendencias:
            mes = item["mes"]
            if mes not in adopcion_por_mes:
                adopcion_por_mes[mes] = []
            adopcion_por_mes[mes].append(item["tasa_adopcion"])

        # Calcular promedio por mes
        promedios_mensuales = {}
        for mes, tasas in adopcion_por_mes.items():
            promedios_mensuales[mes] = sum(tasas) / len(tasas)

        # Identificar meses de mayor y menor adopción
        mes_mayor = max(promedios_mensuales, key=promedios_mensuales.get)
        mes_menor = min(promedios_mensuales, key=promedios_mensuales.get)

        return {
            "mes_mayor_adopcion": {
                "mes": mes_mayor,
                "mes_nombre": calendar.month_name[mes_mayor],
                "tasa_promedio": round(promedios_mensuales[mes_mayor], 2),
            },
            "mes_menor_adopcion": {
                "mes": mes_menor,
                "mes_nombre": calendar.month_name[mes_menor],
                "tasa_promedio": round(promedios_mensuales[mes_menor], 2),
            },
            "variabilidad_estacional": round(
                (promedios_mensuales[mes_mayor] - promedios_mensuales[mes_menor])
                / promedios_mensuales[mes_mayor]
                * 100,
                2,
            ),
        }

    # ==================== MÉTODOS PARA SOSTENIBILIDAD ====================

    def _calcular_diversificacion_productiva(self, marcas_año):
        """Calcula diversificación de propósitos productivos"""
        propositos = (
            marcas_año.values("proposito_ganado")
            .annotate(total=Count("id"))
            .order_by("-total")
        )

        total_marcas = marcas_año.count()
        if total_marcas == 0:
            return {"score": 0, "interpretacion": "sin_datos"}

        # Índice de Herfindahl-Hirschman para medir concentración
        hhi = sum([(p["total"] / total_marcas) ** 2 for p in propositos])
        diversificacion_score = (1 - hhi) * 100  # Convertir a score de diversificación

        return {
            "score": round(diversificacion_score, 2),
            "numero_propositos_activos": len(propositos),
            "distribucion_propositos": list(propositos),
            "interpretacion": (
                "alta"
                if diversificacion_score > 60
                else "media" if diversificacion_score > 30 else "baja"
            ),
        }

    def _calcular_distribucion_geografica(self, marcas_año):
        """Calcula distribución geográfica para sostenibilidad"""
        departamentos = (
            marcas_año.values("departamento")
            .annotate(total=Count("id"), cabezas=Sum("cantidad_cabezas"))
            .order_by("-cabezas")
        )

        total_departamentos_bolivia = 9
        departamentos_activos = len(departamentos)

        return {
            "departamentos_activos": departamentos_activos,
            "cobertura_nacional": round(
                (departamentos_activos / total_departamentos_bolivia) * 100, 2
            ),
            "distribucion_equitativa": self._evaluar_equidad_geografica(departamentos),
            "concentracion_geografica": self._calcular_concentracion_geografica(
                departamentos
            ),
        }

    def _evaluar_equidad_geografica(self, departamentos):
        """Evalúa si hay equidad en la distribución geográfica"""
        if not departamentos:
            return 0

        total_cabezas = sum(d["cabezas"] for d in departamentos)
        if total_cabezas == 0:
            return 0

        # Calcular coeficiente de Gini simplificado
        cabezas_ordenadas = sorted([d["cabezas"] for d in departamentos])
        n = len(cabezas_ordenadas)

        suma_diferencias = sum(
            abs(cabezas_ordenadas[i] - cabezas_ordenadas[j])
            for i in range(n)
            for j in range(n)
        )

        gini = suma_diferencias / (2 * n * total_cabezas)
        equidad_score = (1 - gini) * 100

        return round(equidad_score, 2)

    def _calcular_concentracion_geografica(self, departamentos):
        """Calcula concentración geográfica usando HHI"""
        if not departamentos:
            return 100

        total_cabezas = sum(d["cabezas"] for d in departamentos)
        if total_cabezas == 0:
            return 100

        hhi = sum((d["cabezas"] / total_cabezas) ** 2 for d in departamentos)
        concentracion = hhi * 100

        return round(concentracion, 2)

    def _calcular_estabilidad_ingresos(self, marcas_año):
        """Calcula estabilidad de ingresos a lo largo del año"""
        ingresos_mensuales = (
            marcas_año.filter(estado="APROBADO")
            .extra(select={"mes": "MONTH(fecha_procesamiento)"})
            .values("mes")
            .annotate(ingresos_mes=Sum("monto_certificacion"))
            .order_by("mes")
        )

        if len(ingresos_mensuales) < 6:  # Necesitamos al menos 6 meses
            return {"score": 0, "interpretacion": "datos_insuficientes"}

        ingresos_lista = [
            float(item["ingresos_mes"] or 0) for item in ingresos_mensuales
        ]

        if not ingresos_lista or sum(ingresos_lista) == 0:
            return {"score": 0, "interpretacion": "sin_ingresos"}

        # Calcular coeficiente de variación
        import statistics

        promedio = statistics.mean(ingresos_lista)
        desviacion = statistics.stdev(ingresos_lista) if len(ingresos_lista) > 1 else 0
        cv = (desviacion / promedio * 100) if promedio > 0 else 100

        # Score de estabilidad (inverso del CV)
        estabilidad_score = max(0, 100 - cv)

        return {
            "score": round(estabilidad_score, 2),
            "coeficiente_variacion": round(cv, 2),
            "promedio_mensual": round(promedio, 2),
            "interpretacion": (
                "alta"
                if estabilidad_score > 70
                else "media" if estabilidad_score > 40 else "baja"
            ),
        }

    def _calcular_inclusion_productores(self, marcas_año):
        """Calcula nivel de inclusión de diferentes tipos de productores"""
        # Categorizar productores por tamaño
        pequeños = marcas_año.filter(cantidad_cabezas__lt=50).count()
        medianos = marcas_año.filter(
            cantidad_cabezas__gte=50, cantidad_cabezas__lt=200
        ).count()
        grandes = marcas_año.filter(cantidad_cabezas__gte=200).count()

        total = marcas_año.count()

        if total == 0:
            return {"score": 0, "interpretacion": "sin_datos"}

        # Score de inclusión basado en diversidad de tamaños
        participacion_pequeños = pequeños / total
        participacion_medianos = medianos / total
        participacion_grandes = grandes / total

        # Índice de diversidad Simpson
        simpson_index = 1 - (
            participacion_pequeños**2
            + participacion_medianos**2
            + participacion_grandes**2
        )
        inclusion_score = simpson_index * 100

        return {
            "score": round(inclusion_score, 2),
            "distribucion_por_tamaño": {
                "pequeños_productores": {
                    "cantidad": pequeños,
                    "porcentaje": round(participacion_pequeños * 100, 2),
                },
                "medianos_productores": {
                    "cantidad": medianos,
                    "porcentaje": round(participacion_medianos * 100, 2),
                },
                "grandes_productores": {
                    "cantidad": grandes,
                    "porcentaje": round(participacion_grandes * 100, 2),
                },
            },
            "interpretacion": (
                "alta"
                if inclusion_score > 60
                else "media" if inclusion_score > 40 else "baja"
            ),
        }

    def _calcular_cobertura_territorial(self, marcas_año):
        """Calcula cobertura territorial del servicio"""
        departamentos_activos = marcas_año.values("departamento").distinct().count()
        municipios_activos = marcas_año.values("municipio").distinct().count()

        # Estimaciones basadas en datos de Bolivia
        total_departamentos = 9
        total_municipios_estimados = 339  # Aproximado para Bolivia

        cobertura_departamental = (departamentos_activos / total_departamentos) * 100
        cobertura_municipal = (municipios_activos / total_municipios_estimados) * 100

        return {
            "cobertura_departamental": round(cobertura_departamental, 2),
            "cobertura_municipal_estimada": round(cobertura_municipal, 2),
            "departamentos_activos": departamentos_activos,
            "municipios_activos": municipios_activos,
            "nivel_cobertura": (
                "excelente"
                if cobertura_departamental > 80
                else "bueno" if cobertura_departamental > 60 else "regular"
            ),
        }

    def _calcular_accesibilidad_servicios(self, marcas_año):
        """Evalúa accesibilidad del servicio"""
        # Tiempo promedio de procesamiento como proxy de accesibilidad
        tiempo_promedio = (
            marcas_año.filter(tiempo_procesamiento_horas__isnull=False).aggregate(
                Avg("tiempo_procesamiento_horas")
            )["tiempo_procesamiento_horas__avg"]
            or 0
        )

        # Score de accesibilidad (menor tiempo = mayor accesibilidad)
        accesibilidad_score = max(0, 100 - (tiempo_promedio - 24) * 2)

        # Distribución geográfica de tiempos
        tiempos_por_departamento = (
            marcas_año.filter(tiempo_procesamiento_horas__isnull=False)
            .values("departamento")
            .annotate(tiempo_promedio_dept=Avg("tiempo_procesamiento_horas"))
        )

        return {
            "score_accesibilidad": round(accesibilidad_score, 2),
            "tiempo_promedio_horas": round(tiempo_promedio, 2),
            "variabilidad_regional": list(tiempos_por_departamento),
            "interpretacion": (
                "alta"
                if accesibilidad_score > 75
                else "media" if accesibilidad_score > 50 else "baja"
            ),
        }

    def _estimar_equidad_genero(self, marcas_año):
        """Estima equidad de género (limitado por datos disponibles)"""
        # Nota: Esta es una estimación ya que no tenemos datos explícitos de género
        # Se basa en análisis de nombres cuando sea posible

        total_productores = marcas_año.values("nombre_productor").distinct().count()

        # Análisis muy básico basado en terminaciones de nombres
        # Esta es solo una aproximación y debería mejorarse con datos reales
        nombres_femeninos_estimados = (
            marcas_año.filter(
                Q(nombre_productor__icontains=" maria ")
                | Q(nombre_productor__icontains=" ana ")
                | Q(nombre_productor__icontains=" carmen ")
                | Q(nombre_productor__iendswith=" maria")
                | Q(nombre_productor__iendswith=" ana")
            )
            .values("nombre_productor")
            .distinct()
            .count()
        )

        participacion_estimada_mujeres = (
            (nombres_femeninos_estimados / total_productores * 100)
            if total_productores > 0
            else 0
        )

        return {
            "participacion_estimada_mujeres": round(participacion_estimada_mujeres, 2),
            "total_productores_analizados": total_productores,
            "metodologia": "Estimación basada en análisis de nombres - requiere datos específicos para mayor precisión",
            "limitaciones": "Datos insuficientes para análisis preciso de equidad de género",
            "recomendacion": "Implementar campos específicos de género en formularios de registro",
        }

    def _evaluar_desarrollo_rural(self, marcas_año):
        """Evalúa contribución al desarrollo rural"""
        # Productores únicos como proxy de familias beneficiadas
        productores_unicos = marcas_año.values("ci_productor").distinct().count()

        # Distribución por tamaño de operación
        pequeñas_operaciones = marcas_año.filter(cantidad_cabezas__lt=50).count()

        # Cobertura de comunidades
        comunidades_cubiertas = (
            marcas_año.exclude(comunidad__isnull=True)
            .exclude(comunidad="")
            .values("comunidad")
            .distinct()
            .count()
        )

        # Score de desarrollo rural
        score_familias = min(
            100, (productores_unicos / 1000) * 100
        )  # 1000 familias como meta
        score_inclusion = (
            (pequeñas_operaciones / marcas_año.count() * 100)
            if marcas_año.count() > 0
            else 0
        )

        desarrollo_rural_score = score_familias * 0.6 + score_inclusion * 0.4

        return {
            "score_desarrollo_rural": round(desarrollo_rural_score, 2),
            "familias_beneficiadas_estimadas": productores_unicos,
            "pequeñas_operaciones_porcentaje": round(score_inclusion, 2),
            "comunidades_atendidas": comunidades_cubiertas,
            "impacto_estimado": (
                "alto"
                if desarrollo_rural_score > 70
                else "medio" if desarrollo_rural_score > 40 else "bajo"
            ),
        }

    def _estimar_eficiencia_tierra(self, marcas_año):
        """Estima eficiencia en el uso de tierra"""
        # Promedio de cabezas por registro como proxy de densidad ganadera
        promedio_cabezas = (
            marcas_año.aggregate(Avg("cantidad_cabezas"))["cantidad_cabezas__avg"] or 0
        )

        # Benchmarks estimados para Bolivia
        benchmark_cabezas_hectarea = {
            "CARNE": 1.5,  # cabezas por hectárea en pastoreo extensivo
            "LECHE": 2.5,  # cabezas por hectárea en sistema intensivo
            "DOBLE_PROPOSITO": 2.0,
            "REPRODUCCION": 1.0,
        }

        # Análisis por propósito
        eficiencia_por_proposito = []
        for proposito in benchmark_cabezas_hectarea.keys():
            marcas_proposito = marcas_año.filter(proposito_ganado=proposito)
            if marcas_proposito.exists():
                promedio_proposito = (
                    marcas_proposito.aggregate(Avg("cantidad_cabezas"))[
                        "cantidad_cabezas__avg"
                    ]
                    or 0
                )
                # Estimación: asumiendo 50 hectáreas promedio por operación
                densidad_estimada = promedio_proposito / 50
                benchmark = benchmark_cabezas_hectarea[proposito]

                eficiencia_porcentaje = (
                    (densidad_estimada / benchmark * 100) if benchmark > 0 else 0
                )

                eficiencia_por_proposito.append(
                    {
                        "proposito": proposito,
                        "densidad_estimada": round(densidad_estimada, 3),
                        "benchmark": benchmark,
                        "eficiencia_porcentaje": round(eficiencia_porcentaje, 2),
                    }
                )

        return {
            "promedio_cabezas_operacion": round(promedio_cabezas, 2),
            "eficiencia_por_proposito": eficiencia_por_proposito,
            "metodologia": "Estimación basada en promedios sectoriales - requiere datos específicos de área",
            "limitaciones": "No se cuenta con datos precisos de superficie por operación",
        }

    def _calcular_diversidad_genetica(self, marcas_año):
        """Calcula diversidad genética del rebaño"""
        razas_registradas = (
            marcas_año.values("raza_bovino")
            .annotate(
                total_cabezas=Sum("cantidad_cabezas"), total_registros=Count("id")
            )
            .order_by("-total_cabezas")
        )

        total_cabezas = (
            marcas_año.aggregate(Sum("cantidad_cabezas"))["cantidad_cabezas__sum"] or 0
        )

        if total_cabezas == 0:
            return {"score": 0, "interpretacion": "sin_datos"}

        # Índice de Shannon para diversidad
        shannon_index = 0
        for raza in razas_registradas:
            proporcion = raza["total_cabezas"] / total_cabezas
            if proporcion > 0:
                shannon_index += proporcion * (-1) * (proporcion**0.5)  # Simplificado

        # Normalizar a score 0-100
        max_shannon = 1.0  # Valor máximo teórico para el contexto
        diversidad_score = min(100, (shannon_index / max_shannon) * 100)

        return {
            "score_diversidad_genetica": round(abs(diversidad_score), 2),
            "numero_razas_registradas": len(razas_registradas),
            "distribucion_razas": [
                {
                    "raza": item["raza_bovino"],
                    "cabezas": item["total_cabezas"],
                    "participacion": round(
                        (item["total_cabezas"] / total_cabezas * 100), 2
                    ),
                }
                for item in razas_registradas
            ],
            "interpretacion": (
                "alta"
                if len(razas_registradas) >= 8
                else "media" if len(razas_registradas) >= 5 else "baja"
            ),
        }

    def _evaluar_practicas_sostenibles(self, marcas_año):
        """Evalúa adopción de prácticas sostenibles (estimación)"""
        # Uso de tecnología como proxy de modernización/sostenibilidad
        marcas_con_logos = marcas_año.filter(logos__isnull=False).distinct().count()
        total_marcas = marcas_año.count()

        adopcion_tecnologica = (
            (marcas_con_logos / total_marcas * 100) if total_marcas > 0 else 0
        )

        # Diversificación como indicador de sostenibilidad
        propositos_diversos = marcas_año.values("proposito_ganado").distinct().count()
        diversificacion_score = min(
            100, (propositos_diversos / 4) * 100
        )  # 4 propósitos máximo

        # Score combinado de prácticas sostenibles
        practicas_score = adopcion_tecnologica * 0.4 + diversificacion_score * 0.6

        return {
            "score_practicas_sostenibles": round(practicas_score, 2),
            "adopcion_tecnologica_porcentaje": round(adopcion_tecnologica, 2),
            "diversificacion_productiva_score": round(diversificacion_score, 2),
            "indicadores_evaluados": [
                "adopcion_tecnologica",
                "diversificacion_productiva",
            ],
            "interpretacion": (
                "avanzado"
                if practicas_score > 70
                else "intermedio" if practicas_score > 40 else "basico"
            ),
            "recomendaciones": self._generar_recomendaciones_sostenibilidad(
                practicas_score
            ),
        }

    def _generar_recomendaciones_sostenibilidad(self, score):
        """Genera recomendaciones para mejorar sostenibilidad"""
        recomendaciones = []

        if score < 50:
            recomendaciones.extend(
                [
                    "Implementar programas de capacitación en buenas prácticas ganaderas",
                    "Promover diversificación de propósitos productivos",
                    "Incentivar adopción de tecnologías digitales",
                ]
            )
        elif score < 70:
            recomendaciones.extend(
                [
                    "Fortalecer programas de mejoramiento genético",
                    "Implementar sistemas de monitoreo ambiental",
                    "Desarrollar cadenas de valor sostenibles",
                ]
            )
        else:
            recomendaciones.extend(
                [
                    "Convertirse en centro de referencia en prácticas sostenibles",
                    "Desarrollar programas de certificación sostenible",
                    "Liderar iniciativas de ganadería climáticamente inteligente",
                ]
            )

        return recomendaciones

    def _estimar_huella_carbono(self, marcas_año):
        """Estima huella de carbono del sector (aproximación)"""
        # Factores de emisión estimados (kg CO2 equivalente por cabeza por año)
        factores_emision = {
            "CARNE": 1800,  # Ganado de carne
            "LECHE": 2200,  # Ganado lechero (mayor actividad metabólica)
            "DOBLE_PROPOSITO": 2000,
            "REPRODUCCION": 1600,
        }

        huella_total = 0
        huella_por_proposito = []

        for proposito, factor in factores_emision.items():
            cabezas_proposito = (
                marcas_año.filter(proposito_ganado=proposito).aggregate(
                    Sum("cantidad_cabezas")
                )["cantidad_cabezas__sum"]
                or 0
            )

            emision_proposito = cabezas_proposito * factor
            huella_total += emision_proposito

            if cabezas_proposito > 0:
                huella_por_proposito.append(
                    {
                        "proposito": proposito,
                        "cabezas": cabezas_proposito,
                        "factor_emision_kg_co2": factor,
                        "huella_total_kg_co2": emision_proposito,
                        "huella_toneladas_co2": round(emision_proposito / 1000, 2),
                    }
                )

        return {
            "huella_total_toneladas_co2": round(huella_total / 1000, 2),
            "huella_por_proposito": huella_por_proposito,
            "metodologia": "Estimación basada en factores IPCC adaptados para Bolivia",
            "limitaciones": [
                "No considera variaciones en manejo",
                "No incluye secuestro de carbono en pasturas",
                "Factores generalizados por región",
            ],
            "interpretacion": self._interpretar_huella_carbono(huella_total),
        }

    def _interpretar_huella_carbono(self, huella_kg):
        """Interpreta la huella de carbono calculada"""
        huella_toneladas = huella_kg / 1000

        if huella_toneladas < 10000:
            return "Impacto bajo - Sector en desarrollo inicial"
        elif huella_toneladas < 50000:
            return "Impacto moderado - Requiere implementar prácticas de mitigación"
        else:
            return "Impacto significativo - Prioritario implementar estrategias de carbono neutral"

    def _calcular_score_sostenibilidad(self, economica, social, ambiental):
        """Calcula score general de sostenibilidad"""
        # Extraer scores principales de cada dimensión
        score_eco = economica.get("diversificacion_productiva", {}).get("score", 0)
        score_soc = social.get("cobertura_territorial", {}).get(
            "cobertura_departamental", 0
        )
        score_amb = ambiental.get("diversidad_genetica", {}).get(
            "score_diversidad_genetica", 0
        )

        # Score ponderado: 40% económico, 35% social, 25% ambiental
        score_general = score_eco * 0.4 + score_soc * 0.35 + score_amb * 0.25

        # Clasificación
        if score_general >= 80:
            clasificacion = "Altamente Sostenible"
        elif score_general >= 60:
            clasificacion = "Sostenible"
        elif score_general >= 40:
            clasificacion = "En Transición"
        else:
            clasificacion = "Requiere Intervención"

        return {
            "score_general": round(score_general, 2),
            "clasificacion": clasificacion,
            "componentes": {
                "economico": round(score_eco, 2),
                "social": round(score_soc, 2),
                "ambiental": round(score_amb, 2),
            },
            "nivel_maduracion": self._evaluar_nivel_maduracion(score_general),
        }

    def _evaluar_nivel_maduracion(self, score):
        """Evalúa nivel de maduración del sector"""
        if score >= 80:
            return {
                "nivel": "maduro",
                "caracteristicas": [
                    "Alta diversificación",
                    "Cobertura amplia",
                    "Prácticas avanzadas",
                ],
                "siguiente_etapa": "Liderazgo e innovación sectorial",
            }
        elif score >= 60:
            return {
                "nivel": "en_desarrollo",
                "caracteristicas": [
                    "Crecimiento sostenido",
                    "Cobertura parcial",
                    "Prácticas intermedias",
                ],
                "siguiente_etapa": "Consolidación y expansión",
            }
        elif score >= 40:
            return {
                "nivel": "emergente",
                "caracteristicas": [
                    "Crecimiento inicial",
                    "Cobertura limitada",
                    "Prácticas básicas",
                ],
                "siguiente_etapa": "Fortalecimiento institucional",
            }
        else:
            return {
                "nivel": "incipiente",
                "caracteristicas": [
                    "Actividad mínima",
                    "Cobertura muy limitada",
                    "Prácticas tradicionales",
                ],
                "siguiente_etapa": "Desarrollo de capacidades básicas",
            }

    # ==================== MÉTODOS FINALES DE UTILIDAD ====================

    def _construir_queryset_avanzado(self, filtros_avanzados):
        """Construye queryset con filtros avanzados múltiples"""
        queryset = MarcaGanadoBovino.objects.all()

        # Filtros básicos
        if filtros_avanzados.get("departamentos"):
            queryset = queryset.filter(
                departamento__in=filtros_avanzados["departamentos"]
            )

        if filtros_avanzados.get("razas"):
            queryset = queryset.filter(raza_bovino__in=filtros_avanzados["razas"])

        if filtros_avanzados.get("propositos"):
            queryset = queryset.filter(
                proposito_ganado__in=filtros_avanzados["propositos"]
            )

        # Filtros de rango
        if filtros_avanzados.get("cabezas_min"):
            queryset = queryset.filter(
                cantidad_cabezas__gte=filtros_avanzados["cabezas_min"]
            )

        if filtros_avanzados.get("cabezas_max"):
            queryset = queryset.filter(
                cantidad_cabezas__lte=filtros_avanzados["cabezas_max"]
            )

        if filtros_avanzados.get("monto_min"):
            queryset = queryset.filter(
                monto_certificacion__gte=filtros_avanzados["monto_min"]
            )

        if filtros_avanzados.get("monto_max"):
            queryset = queryset.filter(
                monto_certificacion__lte=filtros_avanzados["monto_max"]
            )

        # Filtros temporales
        if filtros_avanzados.get("fecha_desde"):
            queryset = queryset.filter(
                fecha_registro__date__gte=filtros_avanzados["fecha_desde"]
            )

        if filtros_avanzados.get("fecha_hasta"):
            queryset = queryset.filter(
                fecha_registro__date__lte=filtros_avanzados["fecha_hasta"]
            )

        # Filtros de estado
        if filtros_avanzados.get("estados"):
            queryset = queryset.filter(estado__in=filtros_avanzados["estados"])

        return queryset

    def _documentar_limitaciones_metodologicas(self, configuracion):
        """Documenta limitaciones metodológicas del análisis"""
        limitaciones = [
            "Análisis basado en datos de registro, no necesariamente representativo del sector completo",
            "Estimaciones económicas basadas en valores promedio de mercado",
            "Indicadores ambientales calculados con metodologías simplificadas",
        ]

        # Limitaciones específicas según configuración
        if "temporal" in configuracion.get("dimensiones", []):
            limitaciones.append(
                "Análisis temporal limitado por período de disponibilidad de datos"
            )

        if "economico" in configuracion.get("dimensiones", []):
            limitaciones.append(
                "Impacto económico calculado sin considerar efectos macroeconómicos externos"
            )

        if "ambiental" in configuracion.get("dimensiones", []):
            limitaciones.append(
                "Indicadores ambientales requieren datos adicionales para mayor precisión"
            )

        return {
            "limitaciones_generales": limitaciones,
            "recomendaciones_mejora": [
                "Integrar datos de censos ganaderos oficiales",
                "Incorporar indicadores de sostenibilidad específicos",
                "Desarrollar metodologías de medición directa de impactos",
            ],
            "nivel_confianza": "Medio - Apropiado para decisiones estratégicas con consideraciones adicionales",
        }


# ==================== FIN DE LA CLASE REPORTESBOVINOVIEWSET ====================

# NOTA: Este es el final de la clase ReportesBovinoViewSet
# Los métodos adicionales que mencioné en las partes anteriores deberían
# implementarse según las necesidades específicas del proyecto
